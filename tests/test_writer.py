"""Excel出力モジュールのテスト

Property 26: 出力Excel生成
*For any* 翻訳結果について、出力Excelにはメイン翻訳結果シート、パス別比較シート、
コストレポートシートが含まれ、条件付き書式が適用される

**Feature: excel-translation-api, Property 26: 出力Excel生成**
**Validates: Requirements 12.1-12.7**
"""

import os
import tempfile
from pathlib import Path
from typing import Any, Dict, List

import pandas as pd
import pytest
from hypothesis import given, settings, strategies as st, assume
from openpyxl import load_workbook

from excel_translator.writer import (
    ExcelWriter,
    ExcelWriterError,
    MainSheetColumns,
    PassComparisonColumns,
    PASS_COLORS,
)


# ============================================================================
# テストユーティリティ
# ============================================================================

def create_translation_result(
    text_id: str,
    source_text: str,
    translated_text: str,
    character_id: str = None,
    remarks: str = "",
    alternative: str = "",
    glossary_check: str = "",
) -> Dict[str, Any]:
    """テスト用翻訳結果を作成"""
    return {
        "text_id": text_id,
        "source_text": source_text,
        "translated_text": translated_text,
        "character_id": character_id,
        "remarks": remarks,
        "alternative": alternative,
        "glossary_check": glossary_check,
    }


def create_pass_result(
    text_id: str,
    source_text: str,
    pass_1: str,
    pass_2: str = None,
    pass_2_reason: str = None,
    pass_3: str = None,
    pass_3_reason: str = None,
    pass_4_backtrans: str = None,
    final: str = None,
) -> Dict[str, Any]:
    """テスト用パス別結果を作成"""
    return {
        "text_id": text_id,
        "source_text": source_text,
        "pass_1": pass_1,
        "pass_2": pass_2,
        "pass_2_reason": pass_2_reason,
        "pass_3": pass_3,
        "pass_3_reason": pass_3_reason,
        "pass_4_backtrans": pass_4_backtrans,
        "final": final or pass_1,
    }


def create_cost_report(
    model: str = "claude-3-5-sonnet-20241022",
    passes: List[Dict] = None,
) -> Dict[str, Any]:
    """テスト用コストレポートを作成"""
    if passes is None:
        passes = [
            {
                "pass_name": "1st_pass",
                "input_tokens": 1000,
                "output_tokens": 500,
                "cache_hit_tokens": 100,
                "api_calls": 5,
                "processing_time_ms": 2000,
                "cost_usd": 0.01,
                "modified_rows": 0,
                "total_rows": 10,
                "modification_rate": 0.0,
            }
        ]
    
    total_input = sum(p.get("input_tokens", 0) for p in passes)
    total_output = sum(p.get("output_tokens", 0) for p in passes)
    total_cache = sum(p.get("cache_hit_tokens", 0) for p in passes)
    total_calls = sum(p.get("api_calls", 0) for p in passes)
    total_time = sum(p.get("processing_time_ms", 0) for p in passes)
    total_cost = sum(p.get("cost_usd", 0) for p in passes)
    
    return {
        "model": model,
        "passes": passes,
        "summary": {
            "total_input_tokens": total_input,
            "total_output_tokens": total_output,
            "total_cache_hit_tokens": total_cache,
            "total_api_calls": total_calls,
            "total_processing_time_ms": total_time,
            "total_cost_usd": total_cost,
        },
    }


# ============================================================================
# ユニットテスト
# ============================================================================

class TestExcelWriterMainSheet:
    """メインシート出力のテスト"""
    
    def test_write_main_sheet_basic(self):
        """基本的なメインシート出力"""
        writer = ExcelWriter()
        results = [
            create_translation_result("001", "こんにちは", "Hello"),
            create_translation_result("002", "さようなら", "Goodbye"),
        ]
        
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            output_path = f.name
        
        try:
            writer.write_main_sheet(results, output_path)
            
            # ファイルが作成されたことを確認
            assert Path(output_path).exists()
            
            # 内容を確認
            wb = load_workbook(output_path)
            ws = wb.active
            
            # ヘッダー行を確認
            headers = [cell.value for cell in ws[1]]
            assert "text_id" in headers
            assert "source_text" in headers
            assert "translated_text" in headers
            
            # データ行を確認
            assert ws.cell(row=2, column=1).value == "001"
            assert ws.cell(row=3, column=1).value == "002"
            
            wb.close()
        finally:
            if Path(output_path).exists():
                os.unlink(output_path)
    
    def test_write_main_sheet_with_char_limit(self):
        """文字数制限付きメインシート出力"""
        writer = ExcelWriter(char_limit=10)
        results = [
            create_translation_result("001", "短い", "Short"),
            create_translation_result("002", "長い文章", "This is a very long text"),
        ]
        
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            output_path = f.name
        
        try:
            writer.write_main_sheet(results, output_path)
            
            wb = load_workbook(output_path)
            ws = wb.active
            
            # length_okカラムを確認
            headers = [cell.value for cell in ws[1]]
            length_ok_col = headers.index("length_ok") + 1
            
            # 短いテキストは✅
            assert ws.cell(row=2, column=length_ok_col).value == "✅"
            # 長いテキストは⚠️
            assert ws.cell(row=3, column=length_ok_col).value == "⚠️"
            
            wb.close()
        finally:
            if Path(output_path).exists():
                os.unlink(output_path)
    
    def test_write_main_sheet_from_dataframe(self):
        """DataFrameからのメインシート出力"""
        writer = ExcelWriter()
        df = pd.DataFrame([
            {"text_id": "001", "source_text": "テスト", "translated_text": "Test"},
        ])
        
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            output_path = f.name
        
        try:
            writer.write_main_sheet(df, output_path)
            assert Path(output_path).exists()
        finally:
            if Path(output_path).exists():
                os.unlink(output_path)


class TestExcelWriterPassComparison:
    """パス別比較シート出力のテスト"""
    
    def test_write_pass_comparison_sheet(self):
        """パス別比較シート出力"""
        writer = ExcelWriter()
        
        # まずメインシートを作成
        results = [create_translation_result("001", "テスト", "Test")]
        
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            output_path = f.name
        
        try:
            writer.write_main_sheet(results, output_path)
            
            # パス別比較シートを追加
            pass_results = [
                create_pass_result(
                    "001", "テスト", "Test",
                    pass_2="Test revised",
                    pass_2_reason="文体改善",
                ),
            ]
            writer.write_pass_comparison_sheet(pass_results, output_path)
            
            # 内容を確認
            wb = load_workbook(output_path)
            assert "パス別比較" in wb.sheetnames
            
            ws = wb["パス別比較"]
            headers = [cell.value for cell in ws[1]]
            assert "pass_1" in headers
            assert "pass_2" in headers
            assert "pass_2_reason" in headers
            
            wb.close()
        finally:
            if Path(output_path).exists():
                os.unlink(output_path)
    
    def test_conditional_formatting_applied(self):
        """条件付き書式が適用されることを確認"""
        writer = ExcelWriter()
        
        results = [create_translation_result("001", "テスト", "Test")]
        
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            output_path = f.name
        
        try:
            writer.write_main_sheet(results, output_path)
            
            pass_results = [
                create_pass_result(
                    "001", "テスト", "Test",
                    pass_2="Test revised",
                    pass_2_reason="改善",
                    pass_3="Test final",
                    pass_3_reason="統一",
                ),
            ]
            writer.write_pass_comparison_sheet(pass_results, output_path)
            
            wb = load_workbook(output_path)
            ws = wb["パス別比較"]
            
            # pass_2カラムに青色が適用されているか確認
            headers = [cell.value for cell in ws[1]]
            pass_2_col = headers.index("pass_2") + 1
            
            cell = ws.cell(row=2, column=pass_2_col)
            # 色が適用されていることを確認（青色）
            fill_color = cell.fill.start_color.rgb
            assert fill_color in ["FFCCE5FF", "CCE5FF", "00CCE5FF"], f"青色が適用されるべき（実際: {fill_color}）"
            
            wb.close()
        finally:
            if Path(output_path).exists():
                os.unlink(output_path)


class TestExcelWriterCostReport:
    """コストレポートシート出力のテスト"""
    
    def test_write_cost_report_sheet(self):
        """コストレポートシート出力"""
        writer = ExcelWriter()
        
        results = [create_translation_result("001", "テスト", "Test")]
        
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            output_path = f.name
        
        try:
            writer.write_main_sheet(results, output_path)
            
            cost_report = create_cost_report()
            writer.write_cost_report_sheet(cost_report, output_path)
            
            wb = load_workbook(output_path)
            assert "コストレポート" in wb.sheetnames
            
            ws = wb["コストレポート"]
            # モデル名が含まれていることを確認
            found_model = False
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and "claude" in str(cell.value).lower():
                        found_model = True
                        break
            assert found_model
            
            wb.close()
        finally:
            if Path(output_path).exists():
                os.unlink(output_path)


class TestExcelWriterModeComparison:
    """モード間比較シート出力のテスト"""
    
    def test_write_mode_comparison_sheet(self):
        """モード間比較シート出力"""
        writer = ExcelWriter()
        
        results = [create_translation_result("001", "テスト", "Test")]
        
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            output_path = f.name
        
        try:
            writer.write_main_sheet(results, output_path)
            
            mode_results = {
                "modes": ["draft", "standard"],
                "comparison": [
                    {
                        "text_id": "001",
                        "source_text": "テスト",
                        "draft_result": "Test",
                        "standard_result": "Test improved",
                        "differences": [{"from_mode": "draft", "to_mode": "standard", "changed": True}],
                    }
                ],
                "cost_comparison": {
                    "draft": {"total_cost_usd": 0.01},
                    "standard": {"total_cost_usd": 0.02, "additional_cost_usd": 0.01},
                },
            }
            writer.write_mode_comparison_sheet(mode_results, output_path)
            
            wb = load_workbook(output_path)
            assert "モード間比較" in wb.sheetnames
            
            wb.close()
        finally:
            if Path(output_path).exists():
                os.unlink(output_path)


class TestExcelWriterAllSheets:
    """全シート一括出力のテスト"""
    
    def test_write_all_sheets(self):
        """全シート一括出力"""
        writer = ExcelWriter()
        
        results = [create_translation_result("001", "テスト", "Test")]
        pass_results = [create_pass_result("001", "テスト", "Test")]
        cost_report = create_cost_report()
        mode_results = {
            "modes": ["draft"],
            "comparison": [{"text_id": "001", "source_text": "テスト", "draft_result": "Test", "differences": []}],
            "cost_comparison": {},
        }
        
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            output_path = f.name
        
        try:
            writer.write_all_sheets(
                results=results,
                pass_results=pass_results,
                cost_report=cost_report,
                mode_results=mode_results,
                output_path=output_path,
            )
            
            wb = load_workbook(output_path)
            
            # 全シートが存在することを確認
            assert "翻訳結果" in wb.sheetnames
            assert "パス別比較" in wb.sheetnames
            assert "コストレポート" in wb.sheetnames
            assert "モード間比較" in wb.sheetnames
            
            wb.close()
        finally:
            if Path(output_path).exists():
                os.unlink(output_path)


class TestExcelWriterMultiFile:
    """複数ファイル出力のテスト"""
    
    def test_write_multi_file_integrated(self):
        """複数ファイルを統合出力"""
        writer = ExcelWriter()
        
        file_results = {
            "file1.xlsx": [create_translation_result("001", "テスト1", "Test1")],
            "file2.xlsx": [create_translation_result("002", "テスト2", "Test2")],
        }
        
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            output_path = f.name
        
        try:
            writer.write_multi_file_output(file_results, output_path, separate_files=False)
            
            wb = load_workbook(output_path)
            
            # 各ファイルのシートが存在することを確認
            assert "file1" in wb.sheetnames
            assert "file2" in wb.sheetnames
            
            wb.close()
        finally:
            if Path(output_path).exists():
                os.unlink(output_path)


# ============================================================================
# Property-Based Tests
# ============================================================================

@st.composite
def translation_results_strategy(draw):
    """翻訳結果リストを生成するストラテジー"""
    num_items = draw(st.integers(min_value=1, max_value=20))
    results = []
    
    # Excel互換の文字のみを使用（制御文字を除外）
    safe_jp_chars = "あいうえおかきくけこさしすせそたちつてとなにぬねのはひふへほまみむめもやゆよらりるれろわをん"
    safe_en_chars = "ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz "
    
    for i in range(num_items):
        text_id = f"TEXT_{i+1:03d}"
        source_text = draw(st.text(
            alphabet=st.sampled_from(safe_jp_chars),
            min_size=1,
            max_size=50
        ))
        translated_text = draw(st.text(
            alphabet=st.sampled_from(safe_en_chars),
            min_size=1,
            max_size=100
        ))
        character_id = draw(st.one_of(
            st.none(),
            st.sampled_from(["キャラA", "キャラB", "キャラC"])
        ))
        # remarksも安全な文字のみ
        remarks = draw(st.text(
            alphabet=st.sampled_from(safe_en_chars + safe_jp_chars),
            max_size=20
        ))
        
        results.append(create_translation_result(
            text_id=text_id,
            source_text=source_text if source_text.strip() else "テスト",
            translated_text=translated_text if translated_text.strip() else "Test",
            character_id=character_id,
            remarks=remarks,
        ))
    
    return results


@st.composite
def pass_results_strategy(draw):
    """パス別結果リストを生成するストラテジー"""
    num_items = draw(st.integers(min_value=1, max_value=20))
    results = []
    
    # Excel互換の文字のみを使用
    safe_chars = "ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789 "
    safe_jp_chars = "あいうえおかきくけこさしすせそたちつてとなにぬねのはひふへほまみむめもやゆよらりるれろわをん"
    
    for i in range(num_items):
        text_id = f"TEXT_{i+1:03d}"
        source_text = draw(st.text(alphabet=st.sampled_from(safe_jp_chars), min_size=1, max_size=30))
        pass_1 = draw(st.text(alphabet=st.sampled_from(safe_chars), min_size=1, max_size=50))
        
        # 2nd passは50%の確率で変更あり
        has_pass_2 = draw(st.booleans())
        pass_2 = draw(st.text(alphabet=st.sampled_from(safe_chars), min_size=1, max_size=50)) if has_pass_2 else None
        pass_2_reason = draw(st.text(alphabet=st.sampled_from(safe_chars + safe_jp_chars), max_size=20)) if has_pass_2 else None
        
        # 3rd passは30%の確率で変更あり
        has_pass_3 = draw(st.booleans()) and draw(st.booleans())
        pass_3 = draw(st.text(alphabet=st.sampled_from(safe_chars), min_size=1, max_size=50)) if has_pass_3 else None
        pass_3_reason = draw(st.text(alphabet=st.sampled_from(safe_chars + safe_jp_chars), max_size=20)) if has_pass_3 else None
        
        # 4th passは20%の確率で実行
        has_pass_4 = draw(st.booleans()) and draw(st.booleans()) and draw(st.booleans())
        pass_4_backtrans = draw(st.text(alphabet=st.sampled_from(safe_jp_chars), min_size=1, max_size=50)) if has_pass_4 else None
        
        results.append(create_pass_result(
            text_id=text_id,
            source_text=source_text if source_text.strip() else "テスト",
            pass_1=pass_1 if pass_1.strip() else "Test",
            pass_2=pass_2 if pass_2 and pass_2.strip() else None,
            pass_2_reason=pass_2_reason,
            pass_3=pass_3 if pass_3 and pass_3.strip() else None,
            pass_3_reason=pass_3_reason,
            pass_4_backtrans=pass_4_backtrans if pass_4_backtrans and pass_4_backtrans.strip() else None,
        ))
    
    return results


@st.composite
def cost_report_strategy(draw):
    """コストレポートを生成するストラテジー"""
    num_passes = draw(st.integers(min_value=1, max_value=4))
    pass_names = ["1st_pass", "2nd_pass", "3rd_pass", "4th_pass"][:num_passes]
    
    passes = []
    for pass_name in pass_names:
        passes.append({
            "pass_name": pass_name,
            "input_tokens": draw(st.integers(min_value=100, max_value=10000)),
            "output_tokens": draw(st.integers(min_value=50, max_value=5000)),
            "cache_hit_tokens": draw(st.integers(min_value=0, max_value=2000)),
            "api_calls": draw(st.integers(min_value=1, max_value=100)),
            "processing_time_ms": draw(st.integers(min_value=100, max_value=60000)),
            "cost_usd": draw(st.floats(min_value=0.001, max_value=10.0)),
            "modified_rows": draw(st.integers(min_value=0, max_value=100)),
            "total_rows": draw(st.integers(min_value=1, max_value=100)),
            "modification_rate": draw(st.floats(min_value=0.0, max_value=1.0)),
        })
    
    return create_cost_report(passes=passes)


class TestProperty26ExcelGeneration:
    """
    Property 26: 出力Excel生成
    
    *For any* 翻訳結果について、出力Excelにはメイン翻訳結果シート、パス別比較シート、
    コストレポートシートが含まれ、条件付き書式が適用される
    
    **Feature: excel-translation-api, Property 26: 出力Excel生成**
    **Validates: Requirements 12.1-12.7**
    """
    
    @given(results=translation_results_strategy())
    @settings(max_examples=100, deadline=None)
    def test_main_sheet_contains_required_columns(self, results: List[Dict[str, Any]]):
        """
        メインシートに必須カラムが含まれる
        Requirements 12.1
        """
        assume(len(results) > 0)
        
        writer = ExcelWriter()
        
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            output_path = f.name
        
        try:
            writer.write_main_sheet(results, output_path)
            
            wb = load_workbook(output_path)
            ws = wb.active
            
            headers = [cell.value for cell in ws[1]]
            
            # 必須カラムが存在することを確認
            required_columns = [
                "text_id", "source_text", "translated_text", "character_id",
                "char_count", "length_ok", "glossary_check", "remarks", "alternative"
            ]
            for col in required_columns:
                assert col in headers, f"必須カラム '{col}' がメインシートに含まれるべき"
            
            wb.close()
        finally:
            if Path(output_path).exists():
                os.unlink(output_path)

    
    @given(results=translation_results_strategy())
    @settings(max_examples=100, deadline=None)
    def test_main_sheet_preserves_all_rows(self, results: List[Dict[str, Any]]):
        """
        メインシートに全行が保持される
        Requirements 12.1
        """
        assume(len(results) > 0)
        
        writer = ExcelWriter()
        
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            output_path = f.name
        
        try:
            writer.write_main_sheet(results, output_path)
            
            wb = load_workbook(output_path)
            ws = wb.active
            
            # データ行数を確認（ヘッダー行を除く）
            data_rows = ws.max_row - 1
            assert data_rows == len(results), "全行がメインシートに保持されるべき"
            
            wb.close()
        finally:
            if Path(output_path).exists():
                os.unlink(output_path)
    
    @given(pass_results=pass_results_strategy())
    @settings(max_examples=100, deadline=None)
    def test_pass_comparison_sheet_contains_required_columns(self, pass_results: List[Dict[str, Any]]):
        """
        パス別比較シートに必須カラムが含まれる
        Requirements 12.2
        """
        assume(len(pass_results) > 0)
        
        writer = ExcelWriter()
        
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            output_path = f.name
        
        try:
            # まずメインシートを作成
            main_results = [
                create_translation_result(r["text_id"], r["source_text"], r["pass_1"])
                for r in pass_results
            ]
            writer.write_main_sheet(main_results, output_path)
            
            # パス別比較シートを追加
            writer.write_pass_comparison_sheet(pass_results, output_path)
            
            wb = load_workbook(output_path)
            ws = wb["パス別比較"]
            
            headers = [cell.value for cell in ws[1]]
            
            # 必須カラムが存在することを確認
            required_columns = [
                "text_id", "source_text", "pass_1", "pass_2", "pass_2_reason",
                "pass_3", "pass_3_reason", "pass_4_backtrans", "final"
            ]
            for col in required_columns:
                assert col in headers, f"必須カラム '{col}' がパス別比較シートに含まれるべき"
            
            wb.close()
        finally:
            if Path(output_path).exists():
                os.unlink(output_path)

    
    @given(cost_report=cost_report_strategy())
    @settings(max_examples=100, deadline=None)
    def test_cost_report_sheet_contains_pass_details(self, cost_report: Dict[str, Any]):
        """
        コストレポートシートにパス別詳細が含まれる
        Requirements 12.3
        """
        writer = ExcelWriter()
        
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            output_path = f.name
        
        try:
            # まずメインシートを作成
            main_results = [create_translation_result("001", "テスト", "Test")]
            writer.write_main_sheet(main_results, output_path)
            
            # コストレポートシートを追加
            writer.write_cost_report_sheet(cost_report, output_path)
            
            wb = load_workbook(output_path)
            ws = wb["コストレポート"]
            
            # パス名が含まれていることを確認
            pass_names = [p["pass_name"] for p in cost_report["passes"]]
            found_passes = set()
            
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and str(cell.value) in pass_names:
                        found_passes.add(str(cell.value))
            
            assert found_passes == set(pass_names), "全パスの詳細がコストレポートに含まれるべき"
            
            wb.close()
        finally:
            if Path(output_path).exists():
                os.unlink(output_path)
    
    @given(
        results=translation_results_strategy(),
        char_limit=st.integers(min_value=5, max_value=100),
    )
    @settings(max_examples=100, deadline=None)
    def test_length_ok_column_reflects_char_limit(self, results: List[Dict[str, Any]], char_limit: int):
        """
        length_okカラムが文字数制限を正しく反映する
        Requirements 12.4
        """
        assume(len(results) > 0)
        
        writer = ExcelWriter(char_limit=char_limit)
        
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            output_path = f.name
        
        try:
            writer.write_main_sheet(results, output_path)
            
            wb = load_workbook(output_path)
            ws = wb.active
            
            headers = [cell.value for cell in ws[1]]
            length_ok_col = headers.index("length_ok") + 1
            char_count_col = headers.index("char_count") + 1
            
            # 各行のlength_okが正しいことを確認
            for row_idx in range(2, ws.max_row + 1):
                char_count = ws.cell(row=row_idx, column=char_count_col).value
                length_ok = ws.cell(row=row_idx, column=length_ok_col).value
                
                if char_count is not None:
                    if int(char_count) <= char_limit:
                        assert length_ok == "✅", f"文字数{char_count}は制限{char_limit}以下なので✅であるべき"
                    else:
                        assert length_ok == "⚠️", f"文字数{char_count}は制限{char_limit}を超えているので⚠️であるべき"
            
            wb.close()
        finally:
            if Path(output_path).exists():
                os.unlink(output_path)

    
    @given(pass_results=pass_results_strategy())
    @settings(max_examples=100, deadline=None)
    def test_conditional_formatting_applied_to_modified_cells(self, pass_results: List[Dict[str, Any]]):
        """
        変更があったセルに条件付き書式が適用される
        Requirements 12.7
        """
        assume(len(pass_results) > 0)
        # pass_2に値がある結果が存在することを確認
        has_pass_2 = any(r.get("pass_2") for r in pass_results)
        assume(has_pass_2)
        
        writer = ExcelWriter()
        
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            output_path = f.name
        
        try:
            # まずメインシートを作成
            main_results = [
                create_translation_result(r["text_id"], r["source_text"], r["pass_1"])
                for r in pass_results
            ]
            writer.write_main_sheet(main_results, output_path)
            
            # パス別比較シートを追加
            writer.write_pass_comparison_sheet(pass_results, output_path)
            
            wb = load_workbook(output_path)
            ws = wb["パス別比較"]
            
            headers = [cell.value for cell in ws[1]]
            pass_2_col = headers.index("pass_2") + 1
            
            # pass_2に値がある行で色が適用されているか確認
            for row_idx, result in enumerate(pass_results, start=2):
                if result.get("pass_2") and str(result.get("pass_2")).strip():
                    cell = ws.cell(row=row_idx, column=pass_2_col)
                    # 青色（CCE5FF）が適用されていることを確認
                    fill_color = cell.fill.start_color.rgb
                    assert fill_color in ["FFCCE5FF", "CCE5FF", "00CCE5FF"], \
                        f"pass_2に値がある行には青色が適用されるべき（実際: {fill_color}）"
            
            wb.close()
        finally:
            if Path(output_path).exists():
                os.unlink(output_path)
    
    @given(
        results=translation_results_strategy(),
        pass_results=pass_results_strategy(),
        cost_report=cost_report_strategy(),
    )
    @settings(max_examples=50, deadline=None)
    def test_all_sheets_generated_together(
        self,
        results: List[Dict[str, Any]],
        pass_results: List[Dict[str, Any]],
        cost_report: Dict[str, Any],
    ):
        """
        全シートが一括生成される
        Requirements 12.1-12.3
        """
        assume(len(results) > 0)
        assume(len(pass_results) > 0)
        
        writer = ExcelWriter()
        
        mode_results = {
            "modes": ["draft", "standard"],
            "comparison": [
                {
                    "text_id": r["text_id"],
                    "source_text": r["source_text"],
                    "draft_result": r["translated_text"],
                    "standard_result": r["translated_text"],
                    "differences": [],
                }
                for r in results[:5]  # 最初の5件のみ
            ],
            "cost_comparison": {},
        }
        
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            output_path = f.name
        
        try:
            writer.write_all_sheets(
                results=results,
                pass_results=pass_results,
                cost_report=cost_report,
                mode_results=mode_results,
                output_path=output_path,
            )
            
            wb = load_workbook(output_path)
            
            # 全シートが存在することを確認
            assert "翻訳結果" in wb.sheetnames, "メイン翻訳結果シートが存在するべき"
            assert "パス別比較" in wb.sheetnames, "パス別比較シートが存在するべき"
            assert "コストレポート" in wb.sheetnames, "コストレポートシートが存在するべき"
            assert "モード間比較" in wb.sheetnames, "モード間比較シートが存在するべき"
            
            wb.close()
        finally:
            if Path(output_path).exists():
                os.unlink(output_path)
