"""用語集管理モジュールのテスト"""

import gc
import os
import tempfile
import time
from pathlib import Path
from typing import Dict, List, Optional

import pandas as pd
import pytest
from hypothesis import given, settings, strategies as st
from openpyxl import Workbook

from excel_translator.glossary import (
    GlossaryEntry,
    GlossaryManager,
    GlossaryManagerError,
    GlossaryVerificationResult,
    GLOSSARY_COLUMN_ALIASES,
)


def safe_unlink(file_path: str, max_retries: int = 3) -> None:
    """Windowsでのファイル削除を安全に行う"""
    gc.collect()
    for i in range(max_retries):
        try:
            if os.path.exists(file_path):
                os.unlink(file_path)
            return
        except PermissionError:
            if i < max_retries - 1:
                time.sleep(0.1)


# ============================================================================
# テストユーティリティ
# ============================================================================

def create_glossary_excel(
    entries: List[Dict],
    file_path: str,
    headers: Optional[List[str]] = None
) -> str:
    """テスト用用語集Excelファイルを作成"""
    wb = Workbook()
    ws = wb.active
    
    # デフォルトヘッダー
    if headers is None:
        headers = ["term_source", "term_target", "category", "context_note", "do_not_translate"]
    
    # ヘッダー行
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)
    
    # データ行
    for row_idx, entry in enumerate(entries, start=2):
        for col_idx, header in enumerate(headers, start=1):
            value = entry.get(header, "")
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    wb.save(file_path)
    wb.close()
    return file_path


# ============================================================================
# ユニットテスト
# ============================================================================

class TestGlossaryManagerBasic:
    """GlossaryManager基本機能のテスト"""
    
    def test_load_file_not_found(self):
        """存在しないファイルでエラー"""
        manager = GlossaryManager()
        with pytest.raises(GlossaryManagerError) as exc_info:
            manager.load("nonexistent.xlsx")
        assert "用語集ファイルが見つかりません" in str(exc_info.value)
    
    def test_load_unsupported_format(self):
        """サポートされていないファイル形式でエラー"""
        manager = GlossaryManager()
        with tempfile.NamedTemporaryFile(suffix=".txt", delete=False) as f:
            f.write(b"test")
            f.flush()
            
            with pytest.raises(GlossaryManagerError) as exc_info:
                manager.load(f.name)
            assert "サポートされていないファイル形式" in str(exc_info.value)
        
        safe_unlink(f.name)
    
    def test_load_basic_glossary(self):
        """基本的な用語集の読み込み"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            file_path = f.name
        
        try:
            create_glossary_excel(
                [
                    {"term_source": "勇者", "term_target": "Hero", "category": "キャラ名"},
                    {"term_source": "魔王", "term_target": "Demon Lord", "category": "キャラ名"},
                ],
                file_path
            )
            
            manager = GlossaryManager()
            entries = manager.load(file_path)
            
            assert len(entries) == 2
            assert entries[0].term_source == "勇者"
            assert entries[0].term_target == "Hero"
            assert entries[0].category == "キャラ名"
        finally:
            safe_unlink(file_path)
    
    def test_load_with_alias_headers(self):
        """エイリアスヘッダーでの読み込み"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            file_path = f.name
        
        try:
            create_glossary_excel(
                [
                    {"原文": "勇者", "訳語": "Hero", "カテゴリ": "キャラ名"},
                ],
                file_path,
                headers=["原文", "訳語", "カテゴリ", "備考", "翻訳禁止"]
            )
            
            manager = GlossaryManager()
            entries = manager.load(file_path)
            
            assert len(entries) == 1
            assert entries[0].term_source == "勇者"
            assert entries[0].term_target == "Hero"
        finally:
            safe_unlink(file_path)
    
    def test_load_with_do_not_translate(self):
        """翻訳禁止フラグの読み込み"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            file_path = f.name
        
        try:
            create_glossary_excel(
                [
                    {"term_source": "HP", "term_target": "HP", "do_not_translate": "true"},
                    {"term_source": "MP", "term_target": "MP", "do_not_translate": "yes"},
                    {"term_source": "勇者", "term_target": "Hero", "do_not_translate": "false"},
                ],
                file_path
            )
            
            manager = GlossaryManager()
            entries = manager.load(file_path)
            
            assert len(entries) == 3
            assert entries[0].do_not_translate == True
            assert entries[1].do_not_translate == True
            assert entries[2].do_not_translate == False
        finally:
            safe_unlink(file_path)
    
    def test_load_skips_empty_rows(self):
        """空行のスキップ"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            file_path = f.name
        
        try:
            create_glossary_excel(
                [
                    {"term_source": "勇者", "term_target": "Hero"},
                    {"term_source": "", "term_target": ""},  # 空行
                    {"term_source": "魔王", "term_target": "Demon Lord"},
                ],
                file_path
            )
            
            manager = GlossaryManager()
            entries = manager.load(file_path)
            
            assert len(entries) == 2
        finally:
            safe_unlink(file_path)
    
    def test_filter_by_text_basic(self):
        """基本的なテキストフィルタリング"""
        manager = GlossaryManager()
        entries = [
            GlossaryEntry("勇者", "Hero", "キャラ名"),
            GlossaryEntry("魔王", "Demon Lord", "キャラ名"),
            GlossaryEntry("剣", "Sword", "アイテム"),
        ]
        
        text = "勇者は剣を手に取った。"
        matched = manager.filter_by_text(text, entries)
        
        assert len(matched) == 2
        assert any(e.term_source == "勇者" for e in matched)
        assert any(e.term_source == "剣" for e in matched)
        assert not any(e.term_source == "魔王" for e in matched)
    
    def test_filter_by_text_no_match(self):
        """マッチなしのフィルタリング"""
        manager = GlossaryManager()
        entries = [
            GlossaryEntry("勇者", "Hero", "キャラ名"),
            GlossaryEntry("魔王", "Demon Lord", "キャラ名"),
        ]
        
        text = "村人は畑を耕していた。"
        matched = manager.filter_by_text(text, entries)
        
        assert len(matched) == 0
    
    def test_filter_by_text_empty_text(self):
        """空テキストのフィルタリング"""
        manager = GlossaryManager()
        entries = [
            GlossaryEntry("勇者", "Hero", "キャラ名"),
        ]
        
        matched = manager.filter_by_text("", entries)
        assert len(matched) == 0
    
    def test_verify_usage_consistent(self):
        """用語一貫性検証（一貫している場合）"""
        manager = GlossaryManager()
        entries = [
            GlossaryEntry("勇者", "Hero", "キャラ名"),
        ]
        
        source = "勇者は旅に出た。"
        translated = "The Hero set out on a journey."
        
        results = manager.verify_usage(source, translated, entries)
        
        assert len(results) == 0  # 不一致なし
    
    def test_verify_usage_inconsistent(self):
        """用語一貫性検証（不一致がある場合）"""
        manager = GlossaryManager()
        entries = [
            GlossaryEntry("勇者", "Hero", "キャラ名"),
        ]
        
        source = "勇者は旅に出た。"
        translated = "The Warrior set out on a journey."  # Heroではなく Warrior
        
        results = manager.verify_usage(source, translated, entries)
        
        assert len(results) == 1
        assert results[0].term_source == "勇者"
        assert results[0].is_consistent == False
    
    def test_verify_usage_do_not_translate(self):
        """翻訳禁止用語の検証"""
        manager = GlossaryManager()
        entries = [
            GlossaryEntry("HP", "HP", "ステータス", do_not_translate=True),
        ]
        
        # 正しい場合（HPがそのまま残っている）
        source = "HPが回復した。"
        translated = "HP was restored."
        results = manager.verify_usage(source, translated, entries)
        assert len(results) == 0
        
        # 不正な場合（HPが翻訳されている）
        translated_wrong = "Health Points was restored."
        results = manager.verify_usage(source, translated_wrong, entries)
        assert len(results) == 1
    
    def test_verify_usage_term_not_in_source(self):
        """原文に用語が含まれていない場合"""
        manager = GlossaryManager()
        entries = [
            GlossaryEntry("勇者", "Hero", "キャラ名"),
            GlossaryEntry("魔王", "Demon Lord", "キャラ名"),
        ]
        
        source = "村人は畑を耕していた。"  # 勇者も魔王も含まれていない
        translated = "The villager was plowing the field."
        
        results = manager.verify_usage(source, translated, entries)
        
        assert len(results) == 0  # 原文に用語がないのでチェック対象外
    
    def test_format_for_prompt(self):
        """プロンプト用フォーマット"""
        manager = GlossaryManager()
        entries = [
            GlossaryEntry("勇者", "Hero", "キャラ名", "主人公"),
            GlossaryEntry("HP", "HP", "ステータス", do_not_translate=True),
        ]
        
        formatted = manager.format_for_prompt(entries)
        
        assert "勇者 → Hero" in formatted
        assert "主人公" in formatted
        assert "HP → HP（翻訳禁止）" in formatted
    
    def test_add_entry_and_get_entries(self):
        """エントリの追加と取得"""
        manager = GlossaryManager()
        
        entry = GlossaryEntry("勇者", "Hero", "キャラ名")
        manager.add_entry(entry)
        
        entries = manager.get_entries()
        assert len(entries) == 1
        assert entries[0].term_source == "勇者"
    
    def test_clear(self):
        """用語集のクリア"""
        manager = GlossaryManager()
        manager.add_entry(GlossaryEntry("勇者", "Hero", "キャラ名"))
        
        manager.clear()
        
        assert len(manager.get_entries()) == 0


# ============================================================================
# Property-Based Tests
# ============================================================================

@st.composite
def glossary_entry_strategy(draw):
    """用語集エントリを生成するストラテジー"""
    term_source = draw(st.sampled_from([
        "勇者", "魔王", "剣", "盾", "魔法", "HP", "MP", "村", "城", "森",
        "ドラゴン", "スライム", "ゴブリン", "エルフ", "ドワーフ"
    ]))
    
    term_target = draw(st.sampled_from([
        "Hero", "Demon Lord", "Sword", "Shield", "Magic", "HP", "MP",
        "Village", "Castle", "Forest", "Dragon", "Slime", "Goblin", "Elf", "Dwarf"
    ]))
    
    category = draw(st.sampled_from(["キャラ名", "アイテム", "地名", "ステータス", ""]))
    
    context_note = draw(st.one_of(
        st.none(),
        st.sampled_from(["主人公", "ボス", "武器", "防具", "回復アイテム"])
    ))
    
    do_not_translate = draw(st.booleans())
    
    return GlossaryEntry(
        term_source=term_source,
        term_target=term_target,
        category=category,
        context_note=context_note,
        do_not_translate=do_not_translate
    )


@st.composite
def text_with_terms_strategy(draw, entries: List[GlossaryEntry]):
    """用語を含むテキストを生成するストラテジー"""
    # ベーステキストパーツ
    base_parts = draw(st.lists(
        st.sampled_from([
            "は", "が", "を", "に", "で", "と", "の", "から", "まで",
            "歩いた", "戦った", "見つけた", "手に入れた", "使った",
            "。", "！", "？", "、"
        ]),
        min_size=2,
        max_size=5
    ))
    
    # 用語を選択（0〜全部）
    if entries:
        num_terms = draw(st.integers(min_value=0, max_value=min(3, len(entries))))
        selected_terms = draw(st.lists(
            st.sampled_from([e.term_source for e in entries]),
            min_size=num_terms,
            max_size=num_terms,
            unique=True
        )) if num_terms > 0 else []
    else:
        selected_terms = []
    
    # テキストを組み立て
    result_parts = []
    for i, part in enumerate(base_parts):
        if i < len(selected_terms):
            result_parts.append(selected_terms[i])
        result_parts.append(part)
    
    return "".join(result_parts)


class TestProperty10GlossaryDynamicFiltering:
    """
    Property 10: 用語集動的フィルタリング
    
    *For any* テキストと用語集について、filter_by_textで抽出された用語は全てテキスト内に出現し、
    テキスト内に出現しない用語は含まれない
    
    **Feature: excel-translation-api, Property 10: 用語集動的フィルタリング**
    **Validates: Requirements 3.2, 3.3**
    """
    
    @given(
        entries=st.lists(glossary_entry_strategy(), min_size=1, max_size=10, unique_by=lambda e: e.term_source)
    )
    @settings(max_examples=100, deadline=None)
    def test_filtered_terms_exist_in_text(self, entries: List[GlossaryEntry]):
        """フィルタリングされた用語は全てテキスト内に存在する"""
        manager = GlossaryManager()
        
        # テキストを生成（一部の用語を含む）
        # ランダムに用語を選択してテキストに含める
        import random
        num_to_include = random.randint(0, len(entries))
        included_entries = random.sample(entries, num_to_include)
        
        text_parts = ["テスト文"]
        for entry in included_entries:
            text_parts.append(entry.term_source)
            text_parts.append("を使った")
        text = "".join(text_parts)
        
        # フィルタリング
        matched = manager.filter_by_text(text, entries)
        
        # フィルタリングされた用語は全てテキスト内に存在する
        for entry in matched:
            assert entry.term_source in text, f"'{entry.term_source}' がテキストに含まれていません"
    
    @given(
        entries=st.lists(glossary_entry_strategy(), min_size=1, max_size=10, unique_by=lambda e: e.term_source)
    )
    @settings(max_examples=100, deadline=None)
    def test_non_matching_terms_not_included(self, entries: List[GlossaryEntry]):
        """テキストに含まれない用語はフィルタリング結果に含まれない"""
        manager = GlossaryManager()
        
        # 用語を含まないテキスト
        text = "これは用語を含まないテスト文です。"
        
        # フィルタリング
        matched = manager.filter_by_text(text, entries)
        
        # マッチした用語は全てテキストに含まれている
        for entry in matched:
            assert entry.term_source in text
        
        # テキストに含まれない用語はマッチしていない
        for entry in entries:
            if entry.term_source not in text:
                assert entry not in matched, f"'{entry.term_source}' がテキストに含まれていないのにマッチしています"
    
    @given(
        entries=st.lists(glossary_entry_strategy(), min_size=1, max_size=5, unique_by=lambda e: e.term_source)
    )
    @settings(max_examples=100, deadline=None)
    def test_all_matching_terms_are_included(self, entries: List[GlossaryEntry]):
        """テキストに含まれる全ての用語がフィルタリング結果に含まれる"""
        manager = GlossaryManager()
        
        # 全ての用語を含むテキストを生成
        text = "テスト: " + " ".join([e.term_source for e in entries])
        
        # フィルタリング
        matched = manager.filter_by_text(text, entries)
        
        # 全ての用語がマッチしている
        matched_sources = {e.term_source for e in matched}
        for entry in entries:
            assert entry.term_source in matched_sources, f"'{entry.term_source}' がマッチしていません"


class TestProperty11GlossaryConsistencyVerification:
    """
    Property 11: 用語一貫性検証
    
    *For any* 用語集エントリと翻訳結果について、原文にterm_sourceが含まれる場合、
    翻訳結果にterm_targetが含まれていなければ不一致として検出される
    
    **Feature: excel-translation-api, Property 11: 用語一貫性検証**
    **Validates: Requirements 3.4, 3.5**
    """
    
    @given(
        entries=st.lists(glossary_entry_strategy(), min_size=1, max_size=5, unique_by=lambda e: e.term_source)
    )
    @settings(max_examples=100, deadline=None)
    def test_consistent_translation_no_errors(self, entries: List[GlossaryEntry]):
        """一貫した翻訳では不一致が検出されない"""
        manager = GlossaryManager()
        
        # 原文を生成（全ての用語を含む）
        source = "テスト: " + " ".join([e.term_source for e in entries])
        
        # 一貫した翻訳を生成（全ての訳語を含む）
        translated_parts = ["Test: "]
        for entry in entries:
            if entry.do_not_translate:
                translated_parts.append(entry.term_source)  # 翻訳禁止は原文のまま
            else:
                translated_parts.append(entry.term_target)
            translated_parts.append(" ")
        translated = "".join(translated_parts)
        
        # 検証
        results = manager.verify_usage(source, translated, entries)
        
        # 不一致なし
        assert len(results) == 0, f"不一致が検出されました: {[r.message for r in results]}"
    
    @given(
        entry=glossary_entry_strategy()
    )
    @settings(max_examples=100, deadline=None)
    def test_inconsistent_translation_detected(self, entry: GlossaryEntry):
        """不一致な翻訳が検出される"""
        manager = GlossaryManager()
        
        # 原文（用語を含む）
        source = f"テスト文に{entry.term_source}が含まれています。"
        
        # 不一致な翻訳（訳語を含まない）
        translated = "This is a test sentence without the correct term."
        
        # 検証
        results = manager.verify_usage(source, translated, [entry])
        
        # 不一致が検出される
        assert len(results) == 1
        assert results[0].term_source == entry.term_source
        assert results[0].is_consistent == False
    
    @given(
        entries=st.lists(glossary_entry_strategy(), min_size=1, max_size=5, unique_by=lambda e: e.term_source)
    )
    @settings(max_examples=100, deadline=None)
    def test_terms_not_in_source_not_checked(self, entries: List[GlossaryEntry]):
        """原文に含まれない用語はチェック対象外"""
        manager = GlossaryManager()
        
        # 用語を含まない原文
        source = "これは用語を含まないテスト文です。"
        
        # 訳語も含まない翻訳
        translated = "This is a test sentence without any terms."
        
        # 検証
        results = manager.verify_usage(source, translated, entries)
        
        # 原文に用語がないので不一致は検出されない
        assert len(results) == 0
    
    @given(
        entry=glossary_entry_strategy().filter(lambda e: e.do_not_translate)
    )
    @settings(max_examples=100, deadline=None)
    def test_do_not_translate_preserved(self, entry: GlossaryEntry):
        """翻訳禁止用語が保持されているかチェック"""
        manager = GlossaryManager()
        
        # 原文（翻訳禁止用語を含む）
        source = f"テスト文に{entry.term_source}が含まれています。"
        
        # 正しい翻訳（原文のまま保持）
        translated = f"This test sentence contains {entry.term_source}."
        
        # 検証
        results = manager.verify_usage(source, translated, [entry])
        
        # 不一致なし
        assert len(results) == 0


class TestGlossaryManagerEdgeCases:
    """用語集管理のエッジケーステスト"""
    
    def test_empty_glossary(self):
        """空の用語集"""
        manager = GlossaryManager()
        
        matched = manager.filter_by_text("テスト文", [])
        assert len(matched) == 0
        
        results = manager.verify_usage("原文", "翻訳", [])
        assert len(results) == 0
    
    def test_partial_match(self):
        """部分一致のテスト"""
        manager = GlossaryManager()
        entries = [
            GlossaryEntry("勇者", "Hero", "キャラ名"),
            GlossaryEntry("勇者様", "Lord Hero", "キャラ名"),
        ]
        
        # "勇者様"を含むテキスト
        text = "勇者様が来た。"
        matched = manager.filter_by_text(text, entries)
        
        # 両方マッチする（"勇者"も"勇者様"に含まれる）
        assert len(matched) == 2
    
    def test_context_note_in_format(self):
        """文脈メモのフォーマット"""
        manager = GlossaryManager()
        entries = [
            GlossaryEntry("勇者", "Hero", "キャラ名", "主人公"),
        ]
        
        # 文脈メモあり
        formatted_with = manager.format_for_prompt(entries, include_context=True)
        assert "主人公" in formatted_with
        
        # 文脈メモなし
        formatted_without = manager.format_for_prompt(entries, include_context=False)
        assert "主人公" not in formatted_without
    
    def test_load_with_context_note(self):
        """文脈メモ付きの読み込み"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            file_path = f.name
        
        try:
            create_glossary_excel(
                [
                    {"term_source": "勇者", "term_target": "Hero", "context_note": "主人公のこと"},
                ],
                file_path
            )
            
            manager = GlossaryManager()
            entries = manager.load(file_path)
            
            assert len(entries) == 1
            assert entries[0].context_note == "主人公のこと"
        finally:
            safe_unlink(file_path)
