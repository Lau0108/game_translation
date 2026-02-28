"""Excel解析・前処理モジュールのテスト"""

import gc
import os
import tempfile
import time
from pathlib import Path
from typing import Dict, List, Optional

import pandas as pd
import pytest
from hypothesis import given, settings, strategies as st
from openpyxl import Workbook

from excel_translator.parser import (
    COLUMN_ALIASES,
    ExcelParser,
    ExcelParserError,
    InputRow,
    ParsedExcel,
    ParseSummary,
    POSITION_MAPPING,
)


def safe_unlink(file_path: str, max_retries: int = 3) -> None:
    """Windowsでのファイル削除を安全に行う"""
    gc.collect()
    for i in range(max_retries):
        try:
            if os.path.exists(file_path):
                os.unlink(file_path)
            return
        except PermissionError:
            if i < max_retries - 1:
                time.sleep(0.1)


# ============================================================================
# テストユーティリティ
# ============================================================================

def create_test_excel(
    sheets_data: Dict[str, List[List]],
    file_path: str,
    with_headers: bool = True
) -> str:
    """テスト用Excelファイルを作成"""
    wb = Workbook()
    
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    
    for sheet_name, rows in sheets_data.items():
        ws = wb.create_sheet(title=sheet_name)
        for row_idx, row_data in enumerate(rows, start=1):
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
    
    wb.save(file_path)
    wb.close()
    return file_path


# ============================================================================
# ユニットテスト
# ============================================================================

class TestExcelParserBasic:
    """ExcelParser基本機能のテスト"""
    
    def test_read_excel_file_not_found(self):
        """存在しないファイルでエラー"""
        parser = ExcelParser()
        with pytest.raises(ExcelParserError) as exc_info:
            parser.read_excel("nonexistent.xlsx")
        assert "ファイルが見つかりません" in str(exc_info.value)
    
    def test_read_excel_unsupported_format(self):
        """サポートされていないファイル形式でエラー"""
        parser = ExcelParser()
        with tempfile.NamedTemporaryFile(suffix=".txt", delete=False) as f:
            f.write(b"test")
            f.flush()
            
            with pytest.raises(ExcelParserError) as exc_info:
                parser.read_excel(f.name)
            assert "サポートされていないファイル形式" in str(exc_info.value)
        
        safe_unlink(f.name)
    
    def test_detect_header_with_header(self):
        """ヘッダー行の検出"""
        parser = ExcelParser()
        df = pd.DataFrame([
            ["text_id", "character", "source_text"],
            ["001", "キャラA", "テスト文"],
        ])
        
        header_row = parser.detect_header(df)
        assert header_row == 0
    
    def test_detect_header_without_header(self):
        """ヘッダーなしの検出"""
        parser = ExcelParser()
        df = pd.DataFrame([
            ["001", "キャラA", "テスト文1"],
            ["002", "キャラB", "テスト文2"],
        ])
        
        header_row = parser.detect_header(df)
        assert header_row is None
    
    def test_detect_header_with_alias(self):
        """エイリアスでのヘッダー検出"""
        parser = ExcelParser()
        df = pd.DataFrame([
            ["ID", "話者", "原文"],
            ["001", "キャラA", "テスト文"],
        ])
        
        header_row = parser.detect_header(df)
        assert header_row == 0
    
    def test_map_columns_with_header(self):
        """ヘッダーありのカラムマッピング"""
        parser = ExcelParser()
        df = pd.DataFrame([
            ["text_id", "character", "source_text"],
            ["001", "キャラA", "テスト文"],
        ])
        
        mapped_df, mapping = parser.map_columns(df, header_row=0, file_name="test.xlsx")
        
        assert "text_id" in mapped_df.columns
        assert "character" in mapped_df.columns
        assert "source_text" in mapped_df.columns
        assert len(mapped_df) == 1
    
    def test_map_columns_without_header(self):
        """ヘッダーなしの位置ベースマッピング"""
        parser = ExcelParser()
        df = pd.DataFrame([
            ["001", "キャラA", "テスト文1"],
            ["002", "キャラB", "テスト文2"],
        ])
        
        mapped_df, mapping = parser.map_columns(df, header_row=None, file_name="test.xlsx")
        
        assert "text_id" in mapped_df.columns
        assert "character" in mapped_df.columns
        assert "source_text" in mapped_df.columns
        assert len(mapped_df) == 2
    
    def test_detect_skip_rows_empty_text(self):
        """空テキストのスキップ検出"""
        parser = ExcelParser()
        df = pd.DataFrame({
            "text_id": ["001", "002"],
            "character": ["キャラA", "キャラB"],
            "source_text": ["テスト文", ""],
        })
        
        result = parser.detect_skip_rows(df)
        
        assert result.iloc[0]["_skip"] == False
        assert result.iloc[1]["_skip"] == True
        assert result.iloc[1]["_skip_reason"] == "空テキスト"
    
    def test_detect_skip_rows_variable_only(self):
        """変数のみのスキップ検出"""
        parser = ExcelParser()
        df = pd.DataFrame({
            "text_id": ["001", "002", "003"],
            "character": ["キャラA", "キャラB", "キャラC"],
            "source_text": ["テスト文", "{player_name}", "{var1} {var2}"],
        })
        
        result = parser.detect_skip_rows(df)
        
        assert result.iloc[0]["_skip"] == False
        assert result.iloc[1]["_skip"] == True
        assert result.iloc[1]["_skip_reason"] == "変数のみ"
        assert result.iloc[2]["_skip"] == True
    
    def test_normalize_character_trim(self):
        """キャラ名のトリム"""
        parser = ExcelParser()
        df = pd.DataFrame({
            "text_id": ["001"],
            "character": ["キャラA  "],
            "source_text": ["テスト文"],
        })
        
        result, _ = parser.normalize(df, "test.xlsx", "Sheet1")
        
        assert result.iloc[0]["character"] == "キャラA"


class TestExcelParserIntegration:
    """ExcelParser統合テスト"""
    
    def test_parse_single_sheet(self):
        """単一シートの解析"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            file_path = f.name
        
        try:
            create_test_excel(
                {
                    "Sheet1": [
                        ["text_id", "character", "source_text"],
                        ["001", "キャラA", "テスト文1"],
                        ["002", "キャラB", "テスト文2"],
                    ]
                },
                file_path
            )
            
            parser = ExcelParser()
            result = parser.parse(file_path)
            
            assert result.summary.total_sheets == 1
            assert result.summary.total_rows == 2
            assert len(result.rows) == 2
            assert result.rows[0].text_id == "001"
            assert result.rows[0].character == "キャラA"
            assert result.rows[0].source_text == "テスト文1"
        finally:
            safe_unlink(file_path)
    
    def test_parse_multiple_sheets(self):
        """複数シートの解析"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            file_path = f.name
        
        try:
            create_test_excel(
                {
                    "Sheet1": [
                        ["text_id", "character", "source_text"],
                        ["001", "キャラA", "テスト文1"],
                    ],
                    "Sheet2": [
                        ["text_id", "character", "source_text"],
                        ["002", "キャラB", "テスト文2"],
                    ],
                },
                file_path
            )
            
            parser = ExcelParser()
            result = parser.parse(file_path)
            
            assert result.summary.total_sheets == 2
            assert result.summary.total_rows == 2
            
            sheet_names = {row.sheet_name for row in result.rows}
            assert "Sheet1" in sheet_names
            assert "Sheet2" in sheet_names
        finally:
            safe_unlink(file_path)
    
    def test_parse_without_header(self):
        """ヘッダーなしファイルの解析"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            file_path = f.name
        
        try:
            create_test_excel(
                {
                    "Sheet1": [
                        ["001", "キャラA", "テスト文1"],
                        ["002", "キャラB", "テスト文2"],
                    ]
                },
                file_path,
                with_headers=False
            )
            
            parser = ExcelParser()
            result = parser.parse(file_path)
            
            assert result.summary.total_rows == 2
        finally:
            safe_unlink(file_path)
    
    def test_get_summary_text(self):
        """サマリテキスト生成"""
        parser = ExcelParser()
        parsed = ParsedExcel(
            rows=[
                InputRow("001", "キャラA", "テスト", "Sheet1", "test.xlsx", 1, False, None),
            ],
            summary=ParseSummary(
                total_sheets=1,
                total_rows=1,
                skipped_rows=0,
                skip_reasons={},
                column_mapping={"text_id": "text_id"},
                characters_found={"キャラA"},
            )
        )
        
        summary_text = parser.get_summary_text(parsed)
        
        assert "シート数: 1" in summary_text
        assert "総行数: 1" in summary_text
        assert "キャラA" in summary_text


# ============================================================================
# Property-Based Tests
# ============================================================================

@st.composite
def multi_sheet_data_strategy(draw):
    """複数シートデータを生成するストラテジー"""
    num_sheets = draw(st.integers(min_value=1, max_value=3))
    sheets = {}
    
    for i in range(num_sheets):
        sheet_name = f"Sheet{i+1}"
        header = ["text_id", "character", "source_text"]
        rows = [header]
        
        num_rows = draw(st.integers(min_value=1, max_value=3))
        for j in range(num_rows):
            text_id = f"{i+1}_{j+1:03d}"
            character = draw(st.sampled_from(["キャラA", "キャラB", ""]))
            source_text = f"テスト文{i}_{j}"
            rows.append([text_id, character, source_text])
        
        sheets[sheet_name] = rows
    
    return sheets


class TestProperty1MultipleSheetReading:
    """
    Property 1: 複数シート読み込み
    
    **Feature: excel-translation-api, Property 1: 複数シート読み込み**
    **Validates: Requirements 1.1, 1.10**
    """
    
    @given(sheets_data=multi_sheet_data_strategy())
    @settings(max_examples=100, deadline=None)
    def test_all_sheets_are_read(self, sheets_data: Dict[str, List[List]]):
        """全シートが読み込まれ、シート名が保持される"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            file_path = f.name
        
        try:
            create_test_excel(sheets_data, file_path)
            
            parser = ExcelParser()
            result = parser.parse(file_path)
            
            # 全シートが読み込まれている
            assert result.summary.total_sheets == len(sheets_data)
            
            # 各シート名が識別子として保持されている
            sheet_names_in_result = {row.sheet_name for row in result.rows}
            for sheet_name in sheets_data.keys():
                assert sheet_name in sheet_names_in_result
        finally:
            safe_unlink(file_path)


class TestProperty2ColumnMapping:
    """
    Property 2: カラムマッピング
    
    **Feature: excel-translation-api, Property 2: カラムマッピング**
    **Validates: Requirements 1.2, 1.3, 1.4**
    """
    
    @given(
        header_style=st.sampled_from([
            ["text_id", "character", "source_text"],
            ["ID", "話者", "原文"],
            ["テキストID", "キャラクター", "テキスト"],
            None,
        ]),
        num_rows=st.integers(min_value=1, max_value=3)
    )
    @settings(max_examples=100, deadline=None)
    def test_column_mapping_works(self, header_style: Optional[List[str]], num_rows: int):
        """ヘッダーの有無に関わらずカラムがマッピングされる"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            file_path = f.name
        
        try:
            rows = []
            if header_style is not None:
                rows.append(header_style)
            
            for i in range(num_rows):
                rows.append([f"{i+1:03d}", f"キャラ{i}", f"テスト文{i+1}"])
            
            create_test_excel({"Sheet1": rows}, file_path)
            
            parser = ExcelParser()
            result = parser.parse(file_path)
            
            assert len(result.rows) > 0
            
            for row in result.rows:
                assert row.text_id is not None and row.text_id != ""
                assert row.source_text is not None and row.source_text != ""
        finally:
            safe_unlink(file_path)


class TestProperty5TextIdAutoGeneration:
    """
    Property 5: text_id自動生成
    
    **Feature: excel-translation-api, Property 5: text_id自動生成**
    **Validates: Requirements 1.7**
    """
    
    @given(num_rows=st.integers(min_value=1, max_value=10))
    @settings(max_examples=100, deadline=None)
    def test_text_id_auto_generation_uniqueness(self, num_rows: int):
        """自動生成されたtext_idは一意である"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            file_path = f.name
        
        try:
            rows = [["text_id", "character", "source_text"]]
            for i in range(num_rows):
                rows.append(["", f"キャラ{i}", f"テスト文{i+1}"])
            
            create_test_excel({"Sheet1": rows}, file_path)
            
            parser = ExcelParser()
            result = parser.parse(file_path)
            
            text_ids = [row.text_id for row in result.rows]
            assert len(text_ids) == len(set(text_ids))
            
            for text_id in text_ids:
                assert text_id is not None and text_id.strip() != ""
        finally:
            safe_unlink(file_path)
    
    @given(
        num_sheets=st.integers(min_value=2, max_value=3),
        rows_per_sheet=st.integers(min_value=1, max_value=3)
    )
    @settings(max_examples=100, deadline=None)
    def test_text_id_unique_across_sheets(self, num_sheets: int, rows_per_sheet: int):
        """複数シートでもtext_idは一意"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            file_path = f.name
        
        try:
            sheets = {}
            for s in range(num_sheets):
                rows = [["text_id", "character", "source_text"]]
                for i in range(rows_per_sheet):
                    rows.append(["", f"キャラ{i}", f"テスト文{s}_{i}"])
                sheets[f"Sheet{s+1}"] = rows
            
            create_test_excel(sheets, file_path)
            
            parser = ExcelParser()
            result = parser.parse(file_path)
            
            text_ids = [row.text_id for row in result.rows]
            assert len(text_ids) == len(set(text_ids))
        finally:
            safe_unlink(file_path)


class TestProperty7SkipRowFlag:
    """
    Property 7: 翻訳不要行フラグ
    
    **Feature: excel-translation-api, Property 7: 翻訳不要行フラグ**
    **Validates: Requirements 1.11**
    """
    
    @given(
        empty_indices=st.lists(st.integers(min_value=0, max_value=4), min_size=0, max_size=2, unique=True),
        variable_indices=st.lists(st.integers(min_value=0, max_value=4), min_size=0, max_size=2, unique=True),
    )
    @settings(max_examples=100, deadline=None)
    def test_skip_flag_for_empty_and_variable_only(
        self,
        empty_indices: List[int],
        variable_indices: List[int]
    ):
        """空テキストと変数のみの行にskipフラグが設定される"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            file_path = f.name
        
        try:
            num_rows = 5
            rows = [["text_id", "character", "source_text"]]
            
            expected_skip = set()
            
            for i in range(num_rows):
                text_id = f"{i+1:03d}"
                character = f"キャラ{i}"
                
                if i in empty_indices:
                    source_text = ""
                    expected_skip.add(text_id)
                elif i in variable_indices:
                    source_text = "{player_name}"
                    expected_skip.add(text_id)
                else:
                    source_text = f"通常のテスト文{i+1}"
                
                rows.append([text_id, character, source_text])
            
            create_test_excel({"Sheet1": rows}, file_path)
            
            parser = ExcelParser()
            result = parser.parse(file_path)
            
            for row in result.rows:
                if row.text_id in expected_skip:
                    assert row.skip == True
                else:
                    assert row.skip == False
        finally:
            safe_unlink(file_path)
    
    @given(
        digit_only_count=st.integers(min_value=0, max_value=2),
        normal_count=st.integers(min_value=1, max_value=3)
    )
    @settings(max_examples=100, deadline=None)
    def test_skip_flag_for_digit_only(self, digit_only_count: int, normal_count: int):
        """数字のみの行にskipフラグが設定される"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            file_path = f.name
        
        try:
            rows = [["text_id", "character", "source_text"]]
            digit_only_ids = set()
            
            for i in range(digit_only_count):
                text_id = f"D{i+1:03d}"
                rows.append([text_id, f"キャラ{i}", str(i * 100)])
                digit_only_ids.add(text_id)
            
            for i in range(normal_count):
                text_id = f"N{i+1:03d}"
                rows.append([text_id, f"キャラ{i}", f"通常テキスト{i+1}"])
            
            create_test_excel({"Sheet1": rows}, file_path)
            
            parser = ExcelParser()
            result = parser.parse(file_path)
            
            for row in result.rows:
                if row.text_id in digit_only_ids:
                    assert row.skip == True
                    assert row.skip_reason == "数字のみ"
        finally:
            safe_unlink(file_path)


# ============================================================================
# Tag Protection Tests (Property 8 & 9)
# ============================================================================

from excel_translator.parser import (
    TagProtector,
    TagProtectionResult,
    protect_tags,
    restore_tags,
)


@st.composite
def text_with_variables_strategy(draw):
    """変数を含むテキストを生成するストラテジー"""
    # 変数名のパターン
    var_names = draw(st.lists(
        st.sampled_from([
            "player_name", "item", "count", "0", "1", "name", 
            "character.name", "item.price", "npc_name"
        ]),
        min_size=0,
        max_size=3
    ))
    
    # ベーステキスト
    base_texts = draw(st.lists(
        st.sampled_from([
            "こんにちは", "ありがとう", "さようなら", "テスト", 
            "Hello", "World", "Test", "文章", ""
        ]),
        min_size=1,
        max_size=4
    ))
    
    # テキストを組み立て
    result_parts = []
    for i, base in enumerate(base_texts):
        result_parts.append(base)
        if i < len(var_names):
            result_parts.append(f"{{{var_names[i]}}}")
    
    return "".join(result_parts)


@st.composite
def text_with_control_tags_strategy(draw):
    """制御タグを含むテキストを生成するストラテジー"""
    # 制御タグのパターン
    control_tags = draw(st.lists(
        st.sampled_from([
            "\\n", "\\r", "\\t", "<br>", "<br/>", "<br />",
            "[ruby]", "[/ruby]", "[b]", "[/b]", "[i]", "[/i]",
            "[color=red]", "[/color]", "[size=12]", "[/size]",
            "<color>", "</color>", "&nbsp;", "&lt;", "&gt;",
            "&#123;", "&#x1F600;"
        ]),
        min_size=0,
        max_size=3
    ))
    
    # ベーステキスト
    base_texts = draw(st.lists(
        st.sampled_from([
            "テスト文", "Hello", "こんにちは", "文章", "Text", ""
        ]),
        min_size=1,
        max_size=4
    ))
    
    # テキストを組み立て
    result_parts = []
    for i, base in enumerate(base_texts):
        result_parts.append(base)
        if i < len(control_tags):
            result_parts.append(control_tags[i])
    
    return "".join(result_parts)


@st.composite
def text_with_game_markers_strategy(draw):
    """ゲーム固有マーカーを含むテキストを生成するストラテジー"""
    # ゲーム固有マーカーのパターン
    markers = draw(st.lists(
        st.sampled_from([
            "(heart)", "(star)", "(music)", "(smile)", "(angry)",
            "<<wait>>", "<<shake>>", "<<flash>>",
            "[[effect]]", "[[sound]]", "[[animation]]",
            "%player%", "%item%", "%gold%",
            "@name@", "@title@", "@rank@",
            "#color#", "#size#", "#font#"
        ]),
        min_size=0,
        max_size=3
    ))
    
    # ベーステキスト
    base_texts = draw(st.lists(
        st.sampled_from([
            "大好き", "ありがとう", "すごい", "やった", "Hello", ""
        ]),
        min_size=1,
        max_size=4
    ))
    
    # テキストを組み立て
    result_parts = []
    for i, base in enumerate(base_texts):
        result_parts.append(base)
        if i < len(markers):
            result_parts.append(markers[i])
    
    return "".join(result_parts)


@st.composite
def text_with_mixed_tags_strategy(draw):
    """変数、制御タグ、ゲーム固有マーカーを混合したテキストを生成"""
    # 各種タグ
    variables = draw(st.lists(
        st.sampled_from(["{player}", "{item}", "{count}", "{name}"]),
        min_size=0,
        max_size=2
    ))
    
    control_tags = draw(st.lists(
        st.sampled_from(["\\n", "<br>", "[ruby]", "[/ruby]", "&nbsp;"]),
        min_size=0,
        max_size=2
    ))
    
    markers = draw(st.lists(
        st.sampled_from(["(heart)", "<<wait>>", "[[effect]]", "%gold%"]),
        min_size=0,
        max_size=2
    ))
    
    # ベーステキスト
    base_texts = draw(st.lists(
        st.sampled_from(["こんにちは", "テスト", "Hello", "文章", ""]),
        min_size=1,
        max_size=5
    ))
    
    # 全てのタグを混合
    all_tags = variables + control_tags + markers
    draw(st.randoms()).shuffle(all_tags)
    
    # テキストを組み立て
    result_parts = []
    for i, base in enumerate(base_texts):
        result_parts.append(base)
        if i < len(all_tags):
            result_parts.append(all_tags[i])
    
    return "".join(result_parts)


class TestProperty8TagProtection:
    """
    Property 8: タグ保護
    
    *For any* 変数（{...}）、制御タグ（\\n, <br>等）、ゲーム固有マーカー（(heart)等）を含むテキストについて、
    protect_tags実行後はこれらが全てプレースホルダー（<<VAR_N>>または<<TAG_N>>または<<MARKER_N>>）に置換されている
    
    **Feature: excel-translation-api, Property 8: タグ保護**
    **Validates: Requirements 2.1, 2.2, 2.3**
    """
    
    @given(text=text_with_variables_strategy())
    @settings(max_examples=100, deadline=None)
    def test_variables_are_protected(self, text: str):
        """変数（{...}）がプレースホルダーに置換される"""
        protector = TagProtector()
        result = protector.protect_tags(text)
        
        # 元のテキストに変数が含まれていた場合
        import re
        original_vars = re.findall(r'\{[^}]+\}', text)
        
        if original_vars:
            # プレースホルダーに置換されている
            assert result.var_count == len(original_vars)
            
            # 元の変数は保護済みテキストに含まれていない
            for var in original_vars:
                assert var not in result.protected_text
            
            # プレースホルダーが含まれている
            for i in range(result.var_count):
                assert f"<<VAR_{i}>>" in result.protected_text
            
            # placeholder_mapに元の値が保存されている
            for i in range(result.var_count):
                placeholder = f"<<VAR_{i}>>"
                assert placeholder in result.placeholder_map
                assert result.placeholder_map[placeholder] in original_vars
    
    @given(text=text_with_control_tags_strategy())
    @settings(max_examples=100, deadline=None)
    def test_control_tags_are_protected(self, text: str):
        """制御タグ（\\n, <br>等）がプレースホルダーに置換される"""
        protector = TagProtector()
        result = protector.protect_tags(text)
        
        # 制御タグパターンをチェック
        import re
        control_patterns = [
            r'\\[nrt]',
            r'<br\s*/?>',
            r'</?\s*br\s*>',
            r'\[/?[a-zA-Z_][a-zA-Z0-9_]*\]',
            r'\[/?[a-zA-Z_][a-zA-Z0-9_]*=[^\]]*\]',
            r'</?[a-zA-Z_][a-zA-Z0-9_]*(?:\s+[^>]*)?>',
            r'&[a-zA-Z]+;',
            r'&#\d+;',
            r'&#x[0-9a-fA-F]+;',
        ]
        
        original_tags = []
        for pattern in control_patterns:
            original_tags.extend(re.findall(pattern, text))
        
        if original_tags:
            # タグがプレースホルダーに置換されている
            assert result.tag_count > 0
            
            # プレースホルダーが含まれている
            for i in range(result.tag_count):
                assert f"<<TAG_{i}>>" in result.protected_text
    
    @given(text=text_with_game_markers_strategy())
    @settings(max_examples=100, deadline=None)
    def test_game_markers_are_protected(self, text: str):
        """ゲーム固有マーカー（(heart)等）がプレースホルダーに置換される"""
        protector = TagProtector()
        result = protector.protect_tags(text)
        
        # ゲーム固有マーカーパターンをチェック
        import re
        marker_patterns = [
            r'\([a-zA-Z_][a-zA-Z0-9_]*\)',
            r'<<[a-zA-Z_][a-zA-Z0-9_]*>>',
            r'\[\[[a-zA-Z_][a-zA-Z0-9_]*\]\]',
            r'%[a-zA-Z_][a-zA-Z0-9_]*%',
            r'@[a-zA-Z_][a-zA-Z0-9_]*@',
            r'#[a-zA-Z_][a-zA-Z0-9_]*#',
        ]
        
        original_markers = []
        for pattern in marker_patterns:
            original_markers.extend(re.findall(pattern, text))
        
        if original_markers:
            # マーカーがプレースホルダーに置換されている
            assert result.marker_count > 0
            
            # プレースホルダーが含まれている
            for i in range(result.marker_count):
                assert f"<<MARKER_{i}>>" in result.protected_text
    
    @given(text=text_with_mixed_tags_strategy())
    @settings(max_examples=100, deadline=None)
    def test_all_tag_types_are_protected(self, text: str):
        """全種類のタグが適切にプレースホルダーに置換される"""
        protector = TagProtector()
        result = protector.protect_tags(text)
        
        # 保護済みテキストに元のタグパターンが残っていないことを確認
        import re
        
        # 変数パターン
        remaining_vars = re.findall(r'\{[^}]+\}', result.protected_text)
        assert len(remaining_vars) == 0, f"未保護の変数が残っています: {remaining_vars}"
        
        # プレースホルダーマップの整合性
        total_placeholders = result.var_count + result.tag_count + result.marker_count
        assert len(result.placeholder_map) == total_placeholders


class TestProperty9TagRestorationRoundTrip:
    """
    Property 9: タグ復元ラウンドトリップ
    
    *For any* タグを含むテキストについて、protect_tags → restore_tags を実行すると元のテキストと完全に一致する
    
    **Feature: excel-translation-api, Property 9: タグ復元ラウンドトリップ**
    **Validates: Requirements 2.4, 2.5**
    """
    
    @given(text=text_with_variables_strategy())
    @settings(max_examples=100, deadline=None)
    def test_variable_round_trip(self, text: str):
        """変数を含むテキストのラウンドトリップ"""
        protector = TagProtector()
        
        # 保護
        result = protector.protect_tags(text)
        
        # 復元
        restored = protector.restore_tags(result.protected_text, result.placeholder_map)
        
        # 元のテキストと一致
        assert restored == text
    
    @given(text=text_with_control_tags_strategy())
    @settings(max_examples=100, deadline=None)
    def test_control_tag_round_trip(self, text: str):
        """制御タグを含むテキストのラウンドトリップ"""
        protector = TagProtector()
        
        # 保護
        result = protector.protect_tags(text)
        
        # 復元
        restored = protector.restore_tags(result.protected_text, result.placeholder_map)
        
        # 元のテキストと一致
        assert restored == text
    
    @given(text=text_with_game_markers_strategy())
    @settings(max_examples=100, deadline=None)
    def test_game_marker_round_trip(self, text: str):
        """ゲーム固有マーカーを含むテキストのラウンドトリップ"""
        protector = TagProtector()
        
        # 保護
        result = protector.protect_tags(text)
        
        # 復元
        restored = protector.restore_tags(result.protected_text, result.placeholder_map)
        
        # 元のテキストと一致
        assert restored == text
    
    @given(text=text_with_mixed_tags_strategy())
    @settings(max_examples=100, deadline=None)
    def test_mixed_tags_round_trip(self, text: str):
        """混合タグを含むテキストのラウンドトリップ"""
        protector = TagProtector()
        
        # 保護
        result = protector.protect_tags(text)
        
        # 復元
        restored = protector.restore_tags(result.protected_text, result.placeholder_map)
        
        # 元のテキストと一致
        assert restored == text
    
    @given(text=text_with_mixed_tags_strategy())
    @settings(max_examples=100, deadline=None)
    def test_no_remaining_placeholders_after_restoration(self, text: str):
        """復元後にプレースホルダーが残っていない"""
        protector = TagProtector()
        
        # 保護
        result = protector.protect_tags(text)
        
        # 復元
        restored = protector.restore_tags(result.protected_text, result.placeholder_map)
        
        # プレースホルダーが残っていない
        import re
        remaining = re.findall(r'<<(?:VAR|TAG|MARKER)_\d+>>', restored)
        assert len(remaining) == 0, f"未復元のプレースホルダーが残っています: {remaining}"
    
    @given(text=text_with_mixed_tags_strategy())
    @settings(max_examples=100, deadline=None)
    def test_convenience_functions_round_trip(self, text: str):
        """便利関数（protect_tags, restore_tags）のラウンドトリップ"""
        # 保護
        protected_text, placeholder_map = protect_tags(text)
        
        # 復元
        restored = restore_tags(protected_text, placeholder_map)
        
        # 元のテキストと一致
        assert restored == text


class TestTagProtectorEdgeCases:
    """タグ保護のエッジケーステスト"""
    
    def test_empty_text(self):
        """空テキストの処理"""
        protector = TagProtector()
        result = protector.protect_tags("")
        
        assert result.protected_text == ""
        assert result.placeholder_map == {}
        assert result.var_count == 0
        assert result.tag_count == 0
        assert result.marker_count == 0
    
    def test_text_without_tags(self):
        """タグなしテキストの処理"""
        protector = TagProtector()
        text = "これはタグを含まない普通のテキストです。"
        result = protector.protect_tags(text)
        
        assert result.protected_text == text
        assert result.placeholder_map == {}
        assert result.var_count == 0
        assert result.tag_count == 0
        assert result.marker_count == 0
    
    def test_nested_braces(self):
        """ネストされた波括弧の処理"""
        protector = TagProtector()
        text = "こんにちは{player_name}さん、{item}を{count}個獲得しました。"
        result = protector.protect_tags(text)
        
        assert result.var_count == 3
        assert "{player_name}" not in result.protected_text
        assert "{item}" not in result.protected_text
        assert "{count}" not in result.protected_text
        
        # ラウンドトリップ
        restored = protector.restore_tags(result.protected_text, result.placeholder_map)
        assert restored == text
    
    def test_multiple_same_tags(self):
        """同じタグが複数回出現する場合"""
        protector = TagProtector()
        text = "(heart)大好き(heart)ありがとう(heart)"
        result = protector.protect_tags(text)
        
        assert result.marker_count == 3
        
        # ラウンドトリップ
        restored = protector.restore_tags(result.protected_text, result.placeholder_map)
        assert restored == text
    
    def test_validation_success(self):
        """検証成功ケース"""
        protector = TagProtector()
        text = "{player}さん、(heart)ありがとう\\n"
        result = protector.protect_tags(text)
        restored = protector.restore_tags(result.protected_text, result.placeholder_map)
        
        is_valid, errors = protector.validate_restoration(restored, result.placeholder_map)
        assert is_valid == True
        assert len(errors) == 0
    
    def test_validation_failure_remaining_placeholder(self):
        """検証失敗ケース: プレースホルダーが残っている"""
        protector = TagProtector()
        
        # 意図的にプレースホルダーを残す
        text_with_placeholder = "テスト<<VAR_0>>テスト"
        placeholder_map = {"<<VAR_0>>": "{player}"}
        
        # 復元せずに検証
        is_valid, errors = protector.validate_restoration(text_with_placeholder, placeholder_map)
        assert is_valid == False
        assert len(errors) > 0
