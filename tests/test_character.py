"""キャラクタープロファイル管理モジュールのテスト"""

import gc
import os
import tempfile
import time
from pathlib import Path
from typing import Dict, List, Optional, Set

import pandas as pd
import pytest
from hypothesis import given, settings, strategies as st
from openpyxl import Workbook

from excel_translator.character import (
    CharacterProfile,
    CharacterProfileManager,
    CharacterProfileManagerError,
    CHARACTER_COLUMN_ALIASES,
)


def safe_unlink(file_path: str, max_retries: int = 3) -> None:
    """Windowsでのファイル削除を安全に行う"""
    gc.collect()
    for i in range(max_retries):
        try:
            if os.path.exists(file_path):
                os.unlink(file_path)
            return
        except PermissionError:
            if i < max_retries - 1:
                time.sleep(0.1)


# ============================================================================
# テストユーティリティ
# ============================================================================

def create_character_excel(
    entries: List[Dict],
    file_path: str,
    headers: Optional[List[str]] = None
) -> str:
    """テスト用キャラクターExcelファイルを作成"""
    wb = Workbook()
    ws = wb.active
    
    # デフォルトヘッダー
    if headers is None:
        headers = [
            "character_id", "name_source", "name_target", "personality",
            "speech_style", "speech_examples", "first_person", "second_person",
            "age", "gender", "role"
        ]
    
    # ヘッダー行
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)
    
    # データ行
    for row_idx, entry in enumerate(entries, start=2):
        for col_idx, header in enumerate(headers, start=1):
            value = entry.get(header, "")
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    wb.save(file_path)
    wb.close()
    return file_path


# ============================================================================
# ユニットテスト
# ============================================================================

class TestCharacterProfileManagerBasic:
    """CharacterProfileManager基本機能のテスト"""
    
    def test_load_file_not_found(self):
        """存在しないファイルでエラー"""
        manager = CharacterProfileManager()
        with pytest.raises(CharacterProfileManagerError) as exc_info:
            manager.load("nonexistent.xlsx")
        assert "キャラクターファイルが見つかりません" in str(exc_info.value)
    
    def test_load_unsupported_format(self):
        """サポートされていないファイル形式でエラー"""
        manager = CharacterProfileManager()
        with tempfile.NamedTemporaryFile(suffix=".txt", delete=False) as f:
            f.write(b"test")
            f.flush()
            
            with pytest.raises(CharacterProfileManagerError) as exc_info:
                manager.load(f.name)
            assert "サポートされていないファイル形式" in str(exc_info.value)
        
        safe_unlink(f.name)
    
    def test_load_basic_profile(self):
        """基本的なプロファイルの読み込み"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            file_path = f.name
        
        try:
            create_character_excel(
                [
                    {
                        "character_id": "hero",
                        "name_source": "勇者",
                        "name_target": "Hero",
                        "personality": "勇敢で正義感が強い",
                        "speech_style": "丁寧語",
                        "age": "18",
                        "gender": "男",
                        "role": "主人公",
                    },
                ],
                file_path
            )
            
            manager = CharacterProfileManager()
            profiles = manager.load(file_path)
            
            assert len(profiles) == 1
            assert "hero" in profiles
            assert profiles["hero"].name_source == "勇者"
            assert profiles["hero"].name_target == "Hero"
            assert profiles["hero"].personality == "勇敢で正義感が強い"
            assert profiles["hero"].age == "18"
            assert profiles["hero"].gender == "男"
            assert profiles["hero"].role == "主人公"
        finally:
            safe_unlink(file_path)

    def test_load_with_alias_headers(self):
        """エイリアスヘッダーでの読み込み"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            file_path = f.name
        
        try:
            create_character_excel(
                [
                    {
                        "キャラID": "hero",
                        "原文名": "勇者",
                        "訳語名": "Hero",
                        "性格": "勇敢",
                    },
                ],
                file_path,
                headers=["キャラID", "原文名", "訳語名", "性格", "口調"]
            )
            
            manager = CharacterProfileManager()
            profiles = manager.load(file_path)
            
            assert len(profiles) == 1
            assert "hero" in profiles
            assert profiles["hero"].name_source == "勇者"
            assert profiles["hero"].personality == "勇敢"
        finally:
            safe_unlink(file_path)
    
    def test_load_with_speech_examples_json(self):
        """JSON形式のspeech_examplesの読み込み"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            file_path = f.name
        
        try:
            create_character_excel(
                [
                    {
                        "character_id": "hero",
                        "name_source": "勇者",
                        "name_target": "Hero",
                        "speech_examples": '[["こんにちは", "Hello"], ["ありがとう", "Thank you"]]',
                    },
                ],
                file_path
            )
            
            manager = CharacterProfileManager()
            profiles = manager.load(file_path)
            
            assert len(profiles["hero"].speech_examples) == 2
            assert profiles["hero"].speech_examples[0] == ("こんにちは", "Hello")
            assert profiles["hero"].speech_examples[1] == ("ありがとう", "Thank you")
        finally:
            safe_unlink(file_path)
    
    def test_load_with_speech_examples_pipe(self):
        """パイプ区切りのspeech_examplesの読み込み"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            file_path = f.name
        
        try:
            create_character_excel(
                [
                    {
                        "character_id": "hero",
                        "name_source": "勇者",
                        "name_target": "Hero",
                        "speech_examples": "こんにちは|Hello\nありがとう|Thank you",
                    },
                ],
                file_path
            )
            
            manager = CharacterProfileManager()
            profiles = manager.load(file_path)
            
            assert len(profiles["hero"].speech_examples) == 2
            assert profiles["hero"].speech_examples[0] == ("こんにちは", "Hello")
        finally:
            safe_unlink(file_path)
    
    def test_load_with_speech_examples_semicolon(self):
        """セミコロン区切りのspeech_examplesの読み込み"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            file_path = f.name
        
        try:
            create_character_excel(
                [
                    {
                        "character_id": "hero",
                        "name_source": "勇者",
                        "name_target": "Hero",
                        "speech_examples": "こんにちは|Hello;ありがとう|Thank you",
                    },
                ],
                file_path
            )
            
            manager = CharacterProfileManager()
            profiles = manager.load(file_path)
            
            assert len(profiles["hero"].speech_examples) == 2
        finally:
            safe_unlink(file_path)
    
    def test_load_skips_empty_rows(self):
        """空行のスキップ"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            file_path = f.name
        
        try:
            create_character_excel(
                [
                    {"character_id": "hero", "name_source": "勇者", "name_target": "Hero"},
                    {"character_id": "", "name_source": "", "name_target": ""},  # 空行
                    {"character_id": "villain", "name_source": "魔王", "name_target": "Demon Lord"},
                ],
                file_path
            )
            
            manager = CharacterProfileManager()
            profiles = manager.load(file_path)
            
            assert len(profiles) == 2
        finally:
            safe_unlink(file_path)
    
    def test_get_profiles_for_chunk_basic(self):
        """基本的なチャンクフィルタリング"""
        manager = CharacterProfileManager()
        
        # プロファイルを追加
        manager.add_profile(CharacterProfile(
            character_id="hero",
            name_source="勇者",
            name_target="Hero",
            personality="勇敢",
            speech_style="丁寧語",
            speech_examples=[("こんにちは", "Hello")],
        ))
        manager.add_profile(CharacterProfile(
            character_id="villain",
            name_source="魔王",
            name_target="Demon Lord",
            personality="冷酷",
            speech_style="尊大",
            speech_examples=[("ふはは", "Mwahaha")],
        ))
        manager.add_profile(CharacterProfile(
            character_id="npc",
            name_source="村人",
            name_target="Villager",
            personality="普通",
            speech_style="普通",
        ))
        
        # チャンク内のキャラクターIDセット
        chunk_characters = {"hero", "villain"}
        
        # フィルタリング
        matched = manager.get_profiles_for_chunk(chunk_characters)
        
        assert len(matched) == 2
        matched_ids = {p.character_id for p in matched}
        assert "hero" in matched_ids
        assert "villain" in matched_ids
        assert "npc" not in matched_ids
    
    def test_get_profiles_for_chunk_empty(self):
        """空のチャンクでのフィルタリング"""
        manager = CharacterProfileManager()
        manager.add_profile(CharacterProfile(
            character_id="hero",
            name_source="勇者",
            name_target="Hero",
        ))
        
        matched = manager.get_profiles_for_chunk(set())
        assert len(matched) == 0
    
    def test_get_profiles_for_chunk_no_match(self):
        """マッチなしのフィルタリング"""
        manager = CharacterProfileManager()
        manager.add_profile(CharacterProfile(
            character_id="hero",
            name_source="勇者",
            name_target="Hero",
        ))
        
        matched = manager.get_profiles_for_chunk({"unknown_char"})
        assert len(matched) == 0
    
    def test_format_for_prompt(self):
        """プロンプト用フォーマット"""
        manager = CharacterProfileManager()
        profiles = [
            CharacterProfile(
                character_id="hero",
                name_source="勇者",
                name_target="Hero",
                personality="勇敢で正義感が強い",
                speech_style="丁寧語",
                speech_examples=[("こんにちは", "Hello"), ("ありがとう", "Thank you")],
                first_person="私",
                age="18",
                gender="男",
                role="主人公",
            ),
        ]
        
        formatted = manager.format_for_prompt(profiles)
        
        assert "勇者（Hero）" in formatted
        assert "勇敢で正義感が強い" in formatted
        assert "丁寧語" in formatted
        assert "私" in formatted
        assert "18" in formatted
        assert "男" in formatted
        assert "主人公" in formatted
        assert "こんにちは" in formatted
        assert "Hello" in formatted
    
    def test_format_for_prompt_without_examples(self):
        """例文なしのプロンプト用フォーマット"""
        manager = CharacterProfileManager()
        profiles = [
            CharacterProfile(
                character_id="hero",
                name_source="勇者",
                name_target="Hero",
                speech_examples=[("こんにちは", "Hello")],
            ),
        ]
        
        formatted = manager.format_for_prompt(profiles, include_examples=False)
        
        assert "勇者（Hero）" in formatted
        assert "こんにちは" not in formatted
    
    def test_add_profile_and_get_all(self):
        """プロファイルの追加と取得"""
        manager = CharacterProfileManager()
        
        profile = CharacterProfile(
            character_id="hero",
            name_source="勇者",
            name_target="Hero",
        )
        manager.add_profile(profile)
        
        all_profiles = manager.get_all_profiles()
        assert len(all_profiles) == 1
        assert "hero" in all_profiles
    
    def test_get_profile(self):
        """単一プロファイルの取得"""
        manager = CharacterProfileManager()
        manager.add_profile(CharacterProfile(
            character_id="hero",
            name_source="勇者",
            name_target="Hero",
        ))
        
        profile = manager.get_profile("hero")
        assert profile is not None
        assert profile.name_source == "勇者"
        
        # 存在しないID
        assert manager.get_profile("unknown") is None
    
    def test_clear(self):
        """プロファイルのクリア"""
        manager = CharacterProfileManager()
        manager.add_profile(CharacterProfile(
            character_id="hero",
            name_source="勇者",
            name_target="Hero",
        ))
        
        manager.clear()
        
        assert len(manager.get_all_profiles()) == 0


# ============================================================================
# Property-Based Tests
# ============================================================================

@st.composite
def character_profile_strategy(draw):
    """キャラクタープロファイルを生成するストラテジー"""
    character_id = draw(st.sampled_from([
        "hero", "villain", "npc1", "npc2", "princess", "knight", "mage", "thief"
    ]))
    
    name_source = draw(st.sampled_from([
        "勇者", "魔王", "村人A", "村人B", "姫", "騎士", "魔法使い", "盗賊"
    ]))
    
    name_target = draw(st.sampled_from([
        "Hero", "Demon Lord", "Villager A", "Villager B", "Princess", "Knight", "Mage", "Thief"
    ]))
    
    personality = draw(st.sampled_from([
        "勇敢", "冷酷", "優しい", "臆病", "高貴", "忠実", "知的", "狡猾", ""
    ]))
    
    speech_style = draw(st.sampled_from([
        "丁寧語", "尊大", "普通", "方言", "敬語", "タメ口", ""
    ]))
    
    # speech_examplesを生成
    num_examples = draw(st.integers(min_value=0, max_value=3))
    speech_examples = []
    for _ in range(num_examples):
        source = draw(st.sampled_from([
            "こんにちは", "ありがとう", "さようなら", "助けて", "行くぞ"
        ]))
        target = draw(st.sampled_from([
            "Hello", "Thank you", "Goodbye", "Help me", "Let's go"
        ]))
        speech_examples.append((source, target))
    
    first_person = draw(st.one_of(
        st.none(),
        st.sampled_from(["私", "俺", "僕", "わたくし", "我"])
    ))
    
    second_person = draw(st.one_of(
        st.none(),
        st.sampled_from(["あなた", "君", "お前", "貴様", "そなた"])
    ))
    
    age = draw(st.one_of(
        st.none(),
        st.sampled_from(["10", "18", "25", "40", "100", "不明"])
    ))
    
    gender = draw(st.one_of(
        st.none(),
        st.sampled_from(["男", "女", "不明"])
    ))
    
    role = draw(st.one_of(
        st.none(),
        st.sampled_from(["主人公", "ボス", "仲間", "敵", "NPC"])
    ))
    
    return CharacterProfile(
        character_id=character_id,
        name_source=name_source,
        name_target=name_target,
        personality=personality,
        speech_style=speech_style,
        speech_examples=speech_examples,
        first_person=first_person,
        second_person=second_person,
        age=age,
        gender=gender,
        role=role,
    )


class TestProperty12CharacterProfileDynamicInjection:
    """
    Property 12: キャラプロファイル動的注入
    
    *For any* チャンクと全キャラプロファイルについて、プロンプトに注入されるプロファイルは
    チャンク内に登場するキャラクターのもののみであり、各プロファイルにはspeech_examplesが含まれる
    
    **Feature: excel-translation-api, Property 12: キャラプロファイル動的注入**
    **Validates: Requirements 4.2, 4.3**
    """
    
    @given(
        profiles=st.lists(
            character_profile_strategy(),
            min_size=1,
            max_size=8,
            unique_by=lambda p: p.character_id
        )
    )
    @settings(max_examples=100, deadline=None)
    def test_only_chunk_characters_are_returned(self, profiles: List[CharacterProfile]):
        """チャンク内のキャラクターのプロファイルのみが返される"""
        manager = CharacterProfileManager()
        
        # プロファイルを辞書に変換
        profiles_dict = {p.character_id: p for p in profiles}
        
        # ランダムにチャンク内のキャラクターを選択
        import random
        num_in_chunk = random.randint(0, len(profiles))
        chunk_character_ids = set(random.sample(list(profiles_dict.keys()), num_in_chunk))
        
        # フィルタリング
        matched = manager.get_profiles_for_chunk(chunk_character_ids, profiles_dict)
        
        # 返されたプロファイルは全てチャンク内のキャラクター
        for profile in matched:
            assert profile.character_id in chunk_character_ids, \
                f"'{profile.character_id}' はチャンク内に存在しません"
        
        # チャンク内の全キャラクターのプロファイルが返される
        matched_ids = {p.character_id for p in matched}
        for char_id in chunk_character_ids:
            assert char_id in matched_ids, \
                f"'{char_id}' のプロファイルが返されていません"
    
    @given(
        profiles=st.lists(
            character_profile_strategy(),
            min_size=1,
            max_size=8,
            unique_by=lambda p: p.character_id
        )
    )
    @settings(max_examples=100, deadline=None)
    def test_non_chunk_characters_not_returned(self, profiles: List[CharacterProfile]):
        """チャンク外のキャラクターのプロファイルは返されない"""
        manager = CharacterProfileManager()
        
        # プロファイルを辞書に変換
        profiles_dict = {p.character_id: p for p in profiles}
        
        # 一部のキャラクターのみをチャンクに含める
        all_ids = list(profiles_dict.keys())
        if len(all_ids) > 1:
            # 半分だけチャンクに含める
            chunk_character_ids = set(all_ids[:len(all_ids)//2])
            non_chunk_ids = set(all_ids[len(all_ids)//2:])
        else:
            chunk_character_ids = set(all_ids)
            non_chunk_ids = set()
        
        # フィルタリング
        matched = manager.get_profiles_for_chunk(chunk_character_ids, profiles_dict)
        matched_ids = {p.character_id for p in matched}
        
        # チャンク外のキャラクターは含まれない
        for char_id in non_chunk_ids:
            assert char_id not in matched_ids, \
                f"'{char_id}' はチャンク外なのに返されています"
    
    @given(
        profiles=st.lists(
            character_profile_strategy(),
            min_size=1,
            max_size=5,
            unique_by=lambda p: p.character_id
        )
    )
    @settings(max_examples=100, deadline=None)
    def test_speech_examples_preserved_in_prompt(self, profiles: List[CharacterProfile]):
        """プロンプトフォーマット時にspeech_examplesが保持される"""
        manager = CharacterProfileManager()
        
        # speech_examplesを持つプロファイルのみをフィルタ
        profiles_with_examples = [p for p in profiles if p.speech_examples]
        
        if not profiles_with_examples:
            # speech_examplesがない場合はスキップ
            return
        
        # プロンプトをフォーマット
        formatted = manager.format_for_prompt(profiles_with_examples, include_examples=True)
        
        # 各プロファイルのspeech_examplesがプロンプトに含まれている
        for profile in profiles_with_examples:
            for source, target in profile.speech_examples:
                assert source in formatted, \
                    f"例文の原文 '{source}' がプロンプトに含まれていません"
                assert target in formatted, \
                    f"例文の訳文 '{target}' がプロンプトに含まれていません"
    
    @given(
        profiles=st.lists(
            character_profile_strategy(),
            min_size=1,
            max_size=5,
            unique_by=lambda p: p.character_id
        )
    )
    @settings(max_examples=100, deadline=None)
    def test_empty_chunk_returns_empty_list(self, profiles: List[CharacterProfile]):
        """空のチャンクでは空のリストが返される"""
        manager = CharacterProfileManager()
        
        profiles_dict = {p.character_id: p for p in profiles}
        
        # 空のチャンク
        matched = manager.get_profiles_for_chunk(set(), profiles_dict)
        
        assert len(matched) == 0
    
    @given(
        profiles=st.lists(
            character_profile_strategy(),
            min_size=1,
            max_size=5,
            unique_by=lambda p: p.character_id
        )
    )
    @settings(max_examples=100, deadline=None)
    def test_unknown_character_ids_ignored(self, profiles: List[CharacterProfile]):
        """存在しないキャラクターIDは無視される"""
        manager = CharacterProfileManager()
        
        profiles_dict = {p.character_id: p for p in profiles}
        
        # 存在しないIDを含むチャンク
        chunk_character_ids = {"unknown_id_1", "unknown_id_2"}
        
        # フィルタリング
        matched = manager.get_profiles_for_chunk(chunk_character_ids, profiles_dict)
        
        # 存在しないIDは無視される
        assert len(matched) == 0


class TestCharacterProfileManagerEdgeCases:
    """キャラクタープロファイル管理のエッジケーステスト"""
    
    def test_empty_profiles(self):
        """空のプロファイル辞書"""
        manager = CharacterProfileManager()
        
        matched = manager.get_profiles_for_chunk({"hero"}, {})
        assert len(matched) == 0
    
    def test_single_example_pipe_format(self):
        """単一の例文（パイプ区切り）"""
        manager = CharacterProfileManager()
        examples = manager._parse_speech_examples("こんにちは|Hello")
        
        assert len(examples) == 1
        assert examples[0] == ("こんにちは", "Hello")
    
    def test_empty_speech_examples(self):
        """空のspeech_examples"""
        manager = CharacterProfileManager()
        
        assert manager._parse_speech_examples("") == []
        assert manager._parse_speech_examples(None) == []
        assert manager._parse_speech_examples("nan") == []
    
    def test_format_empty_profiles(self):
        """空のプロファイルリストのフォーマット"""
        manager = CharacterProfileManager()
        
        formatted = manager.format_for_prompt([])
        assert formatted == ""
    
    def test_profile_with_all_optional_fields_none(self):
        """全てのオプションフィールドがNoneのプロファイル"""
        profile = CharacterProfile(
            character_id="test",
            name_source="テスト",
            name_target="Test",
        )
        
        manager = CharacterProfileManager()
        formatted = manager.format_for_prompt([profile])
        
        assert "テスト（Test）" in formatted
        # オプションフィールドは含まれない
        assert "性格:" not in formatted
        assert "口調:" not in formatted
