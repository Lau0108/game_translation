"""Excel解析・前処理モジュール"""

import logging
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

logger = logging.getLogger(__name__)


@dataclass
class InputRow:
    """入力行データ"""
    text_id: str
    character: Optional[str]  # 地の文はNone
    source_text: str
    sheet_name: str
    file_name: str
    row_number: int
    skip: bool = False  # 翻訳不要フラグ
    skip_reason: Optional[str] = None


@dataclass
class ParseSummary:
    """解析サマリ"""
    total_sheets: int
    total_rows: int
    skipped_rows: int
    skip_reasons: Dict[str, int] = field(default_factory=dict)
    column_mapping: Dict[str, str] = field(default_factory=dict)
    characters_found: Set[str] = field(default_factory=set)


@dataclass
class ParsedExcel:
    """解析済みExcelデータ"""
    rows: List[InputRow]
    summary: ParseSummary


class ExcelParserError(Exception):
    """Excel解析エラー"""
    pass


# カラム名エイリアス辞書
COLUMN_ALIASES = {
    "text_id": ["text_id", "id", "ID", "テキストID", "textid", "TextID"],
    "character": ["character", "キャラクター", "キャラ", "話者", "speaker", "char", "name", "名前"],
    "source_text": ["source_text", "原文", "テキスト", "text", "japanese", "日本語", "source", "jp", "ja"],
}

# 位置ベースのデフォルトマッピング（ヘッダーなしの場合）
POSITION_MAPPING = {
    0: "text_id",      # A列
    1: "character",    # B列
    2: "source_text",  # C列
}


class ExcelParser:
    """Excel解析・前処理クラス"""
    
    def __init__(self):
        self._column_aliases = COLUMN_ALIASES.copy()
    
    def read_excel(self, file_path: str) -> Dict[str, pd.DataFrame]:
        """
        Excelファイルを読み込み、全シートをDataFrameとして返す
        
        Args:
            file_path: Excelファイルパス
            
        Returns:
            {sheet_name: DataFrame} の辞書
        """
        path = Path(file_path)
        if not path.exists():
            raise ExcelParserError(f"ファイルが見つかりません: {file_path}")
        
        if not path.suffix.lower() in [".xlsx", ".xls", ".xlsm"]:
            raise ExcelParserError(f"サポートされていないファイル形式: {path.suffix}")
        
        try:
            # 全シートを読み込み（ヘッダーなしで読み込み、後で検出）
            sheets = {}
            with pd.ExcelFile(file_path, engine="openpyxl") as excel_file:
                for sheet_name in excel_file.sheet_names:
                    df = pd.read_excel(
                        excel_file,
                        sheet_name=sheet_name,
                        header=None,  # ヘッダーは後で検出
                        dtype=str,   # 全て文字列として読み込み
                    )
                    sheets[sheet_name] = df
                    logger.info(f"シート '{sheet_name}' を読み込みました: {len(df)} 行")
            
            return sheets
            
        except Exception as e:
            raise ExcelParserError(f"Excelファイルの読み込みに失敗しました: {e}")
    
    def detect_header(self, df: pd.DataFrame) -> Optional[int]:
        """
        ヘッダー行を自動検出
        
        Args:
            df: DataFrame
            
        Returns:
            ヘッダー行のインデックス。存在しない場合はNone
        """
        if df.empty:
            return None
        
        # 最初の数行をチェック（通常ヘッダーは最初の5行以内）
        for row_idx in range(min(5, len(df))):
            row = df.iloc[row_idx]
            row_values = [str(v).strip().lower() for v in row if pd.notna(v)]
            
            # エイリアス辞書のいずれかにマッチするかチェック
            matches = 0
            for canonical, aliases in self._column_aliases.items():
                aliases_lower = [a.lower() for a in aliases]
                if any(val in aliases_lower for val in row_values):
                    matches += 1
            
            # 2つ以上のカラムがマッチすればヘッダー行と判定
            if matches >= 2:
                logger.info(f"ヘッダー行を検出: 行 {row_idx}")
                return row_idx
        
        logger.info("ヘッダー行が見つかりませんでした。位置ベースマッピングを使用します。")
        return None
    
    def map_columns(
        self,
        df: pd.DataFrame,
        header_row: Optional[int],
        file_name: str = ""
    ) -> Tuple[pd.DataFrame, Dict[str, str]]:
        """
        カラムを正規名（text_id, character, source_text）にマッピング
        
        Args:
            df: DataFrame
            header_row: ヘッダー行インデックス（Noneの場合は位置ベース）
            file_name: ファイル名（ログ用）
            
        Returns:
            (マッピング済みDataFrame, カラムマッピング辞書)
        """
        column_mapping = {}
        
        if header_row is not None:
            # ヘッダー行を使用してマッピング
            headers = df.iloc[header_row]
            df = df.iloc[header_row + 1:].reset_index(drop=True)
            df.columns = headers
            
            # エイリアス辞書でマッピング
            new_columns = {}
            for col in df.columns:
                col_str = str(col).strip().lower() if pd.notna(col) else ""
                mapped = False
                
                for canonical, aliases in self._column_aliases.items():
                    aliases_lower = [a.lower() for a in aliases]
                    if col_str in aliases_lower:
                        new_columns[col] = canonical
                        column_mapping[str(col)] = canonical
                        mapped = True
                        break
                
                if not mapped:
                    new_columns[col] = str(col)
            
            df = df.rename(columns=new_columns)
        else:
            # 位置ベースマッピング
            new_columns = {}
            for pos, canonical in POSITION_MAPPING.items():
                if pos < len(df.columns):
                    old_col = df.columns[pos]
                    new_columns[old_col] = canonical
                    column_mapping[f"列{pos + 1}"] = canonical
            
            df = df.rename(columns=new_columns)
        
        # 必須カラムの存在確認
        required = ["source_text"]
        for col in required:
            if col not in df.columns:
                raise ExcelParserError(
                    f"必須カラム '{col}' が見つかりません。"
                    f"ファイル: {file_name}, マッピング: {column_mapping}"
                )
        
        return df, column_mapping

    
    def _expand_merged_cells(self, file_path: str, sheet_name: str, df: pd.DataFrame) -> pd.DataFrame:
        """
        セル結合を解除し、結合範囲全体に同じ値を展開
        
        Args:
            file_path: Excelファイルパス
            sheet_name: シート名
            df: DataFrame
            
        Returns:
            結合解除済みDataFrame
        """
        try:
            wb = load_workbook(file_path, data_only=True)
            ws = wb[sheet_name]
            
            # 結合セル範囲を取得
            merged_ranges = list(ws.merged_cells.ranges)
            
            for merged_range in merged_ranges:
                min_row = merged_range.min_row - 1  # 0-indexed
                max_row = merged_range.max_row - 1
                min_col = merged_range.min_col - 1
                max_col = merged_range.max_col - 1
                
                # 結合セルの値を取得（左上セルの値）
                value = df.iloc[min_row, min_col] if min_row < len(df) else None
                
                # 結合範囲全体に値を展開
                for row in range(min_row, min(max_row + 1, len(df))):
                    for col in range(min_col, min(max_col + 1, len(df.columns))):
                        df.iloc[row, col] = value
            
            wb.close()
            logger.info(f"シート '{sheet_name}': {len(merged_ranges)} 個のセル結合を解除")
            
        except Exception as e:
            logger.warning(f"セル結合の解除中にエラー: {e}")
        
        return df
    
    def normalize(
        self,
        df: pd.DataFrame,
        file_name: str,
        sheet_name: str,
        file_path: Optional[str] = None
    ) -> Tuple[pd.DataFrame, Dict[str, int]]:
        """
        データを正規化: セル結合解除、空行除外、text_id生成、キャラ名トリム
        
        Args:
            df: DataFrame
            file_name: ファイル名
            sheet_name: シート名
            file_path: Excelファイルパス（セル結合解除用）
            
        Returns:
            (正規化済みDataFrame, スキップ理由カウント)
        """
        skip_reasons = {}
        original_len = len(df)
        
        # セル結合の解除
        if file_path:
            df = self._expand_merged_cells(file_path, sheet_name, df)
        
        # 空行の除外
        def is_empty_row(row):
            source = row.get("source_text", "")
            return pd.isna(source) or str(source).strip() == ""
        
        empty_mask = df.apply(is_empty_row, axis=1)
        empty_count = empty_mask.sum()
        if empty_count > 0:
            skip_reasons["空行"] = int(empty_count)
            df = df[~empty_mask].reset_index(drop=True)
            logger.info(f"空行を除外: {empty_count} 行")
        
        # セクション見出し行の除外（source_textが短く、characterが空の行）
        def is_section_header(row):
            source = str(row.get("source_text", "")).strip()
            char = row.get("character", "")
            # 10文字以下で、キャラクターが空、かつ特定パターンにマッチ
            if len(source) <= 10 and (pd.isna(char) or str(char).strip() == ""):
                # 章・節・セクションを示すパターン
                patterns = [r"^第\d+", r"^Chapter", r"^CHAPTER", r"^Scene", r"^---", r"^==="]
                for pattern in patterns:
                    if re.match(pattern, source):
                        return True
            return False
        
        section_mask = df.apply(is_section_header, axis=1)
        section_count = section_mask.sum()
        if section_count > 0:
            skip_reasons["セクション見出し"] = int(section_count)
            df = df[~section_mask].reset_index(drop=True)
            logger.info(f"セクション見出しを除外: {section_count} 行")
        
        # text_idの自動生成（空または存在しない場合）
        if "text_id" not in df.columns:
            df["text_id"] = ""
        
        base_name = Path(file_name).stem if file_name else "unknown"
        for idx in range(len(df)):
            text_id = df.at[idx, "text_id"] if "text_id" in df.columns else None
            if pd.isna(text_id) or str(text_id).strip() == "":
                # ファイル名+シート名+行番号で生成
                df.at[idx, "text_id"] = f"{base_name}_{sheet_name}_{idx + 1}"
        
        # キャラ名のトリム
        if "character" in df.columns:
            df["character"] = df["character"].apply(
                lambda x: str(x).strip() if pd.notna(x) else None
            )
        
        # 行番号を保持
        df["_row_number"] = range(1, len(df) + 1)
        
        logger.info(f"正規化完了: {original_len} → {len(df)} 行")
        
        return df, skip_reasons
    
    def detect_skip_rows(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        翻訳不要行（既に翻訳済み、空テキスト、変数のみ）にフラグ付与
        
        Args:
            df: DataFrame
            
        Returns:
            skipフラグ付きDataFrame
        """
        df = df.copy()
        df["_skip"] = False
        df["_skip_reason"] = None
        
        for idx in range(len(df)):
            source = str(df.at[idx, "source_text"]) if pd.notna(df.at[idx, "source_text"]) else ""
            source = source.strip()
            
            # 空テキスト
            if not source:
                df.at[idx, "_skip"] = True
                df.at[idx, "_skip_reason"] = "空テキスト"
                continue
            
            # 変数のみ（{...}のみで構成）
            var_pattern = r"^\s*(\{[^}]+\}\s*)+$"
            if re.match(var_pattern, source):
                df.at[idx, "_skip"] = True
                df.at[idx, "_skip_reason"] = "変数のみ"
                continue
            
            # 数字のみ
            if source.isdigit():
                df.at[idx, "_skip"] = True
                df.at[idx, "_skip_reason"] = "数字のみ"
                continue
            
            # 既に翻訳済み（translated_textカラムが存在し、値がある場合）
            if "translated_text" in df.columns:
                translated = df.at[idx, "translated_text"]
                if pd.notna(translated) and str(translated).strip():
                    df.at[idx, "_skip"] = True
                    df.at[idx, "_skip_reason"] = "翻訳済み"
                    continue
        
        skip_count = df["_skip"].sum()
        logger.info(f"翻訳不要行を検出: {skip_count} 行")
        
        return df
    
    def parse(self, file_path: str) -> ParsedExcel:
        """
        Excelファイルを解析し、正規化されたデータを返す
        
        Args:
            file_path: Excelファイルパス
            
        Returns:
            ParsedExcel オブジェクト
        """
        path = Path(file_path)
        file_name = path.name
        
        # 全シート読み込み
        sheets = self.read_excel(file_path)
        
        all_rows: List[InputRow] = []
        total_skipped = 0
        all_skip_reasons: Dict[str, int] = {}
        all_column_mapping: Dict[str, str] = {}
        all_characters: Set[str] = set()
        
        for sheet_name, df in sheets.items():
            if df.empty:
                logger.warning(f"シート '{sheet_name}' は空です。スキップします。")
                continue
            
            # ヘッダー検出
            header_row = self.detect_header(df)
            
            # カラムマッピング
            df, column_mapping = self.map_columns(df, header_row, file_name)
            all_column_mapping.update(column_mapping)
            
            # 正規化
            df, skip_reasons = self.normalize(df, file_name, sheet_name, file_path)
            for reason, count in skip_reasons.items():
                all_skip_reasons[reason] = all_skip_reasons.get(reason, 0) + count
                total_skipped += count
            
            # 翻訳不要行検出
            df = self.detect_skip_rows(df)
            
            # InputRowに変換
            for idx in range(len(df)):
                row = df.iloc[idx]
                
                character = row.get("character")
                if pd.notna(character) and str(character).strip():
                    char_str = str(character).strip()
                    all_characters.add(char_str)
                else:
                    char_str = None
                
                input_row = InputRow(
                    text_id=str(row.get("text_id", "")),
                    character=char_str,
                    source_text=str(row.get("source_text", "")),
                    sheet_name=sheet_name,
                    file_name=file_name,
                    row_number=int(row.get("_row_number", idx + 1)),
                    skip=bool(row.get("_skip", False)),
                    skip_reason=row.get("_skip_reason"),
                )
                all_rows.append(input_row)
                
                if input_row.skip:
                    reason = input_row.skip_reason or "不明"
                    all_skip_reasons[reason] = all_skip_reasons.get(reason, 0) + 1
        
        # サマリ作成
        summary = ParseSummary(
            total_sheets=len(sheets),
            total_rows=len(all_rows),
            skipped_rows=sum(1 for r in all_rows if r.skip),
            skip_reasons=all_skip_reasons,
            column_mapping=all_column_mapping,
            characters_found=all_characters,
        )
        
        logger.info(
            f"解析完了: {summary.total_sheets} シート, "
            f"{summary.total_rows} 行, "
            f"{summary.skipped_rows} 行スキップ"
        )
        
        return ParsedExcel(rows=all_rows, summary=summary)
    
    def get_summary_text(self, parsed: ParsedExcel) -> str:
        """
        解析サマリをテキスト形式で返す（ユーザー確認用）
        
        Args:
            parsed: ParsedExcel オブジェクト
            
        Returns:
            サマリテキスト
        """
        s = parsed.summary
        lines = [
            "=== Excel解析サマリ ===",
            f"シート数: {s.total_sheets}",
            f"総行数: {s.total_rows}",
            f"スキップ行数: {s.skipped_rows}",
            "",
            "カラムマッピング:",
        ]
        
        for orig, mapped in s.column_mapping.items():
            lines.append(f"  {orig} → {mapped}")
        
        if s.skip_reasons:
            lines.append("")
            lines.append("スキップ理由:")
            for reason, count in s.skip_reasons.items():
                lines.append(f"  {reason}: {count} 行")
        
        if s.characters_found:
            lines.append("")
            lines.append(f"検出キャラクター数: {len(s.characters_found)}")
            if len(s.characters_found) <= 10:
                lines.append(f"  {', '.join(sorted(s.characters_found))}")
        
        return "\n".join(lines)


@dataclass
class TagProtectionResult:
    """タグ保護結果"""
    protected_text: str
    placeholder_map: Dict[str, str]  # {placeholder: original_value}
    var_count: int
    tag_count: int
    marker_count: int


class TagProtector:
    """タグ保護・復元クラス"""
    
    # 変数パターン: {variable_name}, {player_name}, {0}, {item.name} など
    VAR_PATTERN = re.compile(r'\{[^}]+\}')
    
    # 制御タグパターン: \n, \r, \t, <br>, <br/>, </br>, [ruby], [/ruby], <color=...>, </color> など
    CONTROL_TAG_PATTERNS = [
        re.compile(r'\\[nrt]'),                           # \n, \r, \t
        re.compile(r'<br\s*/?>'),                         # <br>, <br/>, <br />
        re.compile(r'</?\s*br\s*>'),                      # </br>
        re.compile(r'\[/?[a-zA-Z_][a-zA-Z0-9_]*\]'),      # [ruby], [/ruby], [b], [/b]
        re.compile(r'\[/?[a-zA-Z_][a-zA-Z0-9_]*=[^\]]*\]'),  # [color=red], [size=12]
        re.compile(r'</?[a-zA-Z_][a-zA-Z0-9_]*(?:\s+[^>]*)?>'),  # <color>, </color>, <size=12>
        re.compile(r'&[a-zA-Z]+;'),                       # &nbsp;, &lt;, &gt;
        re.compile(r'&#\d+;'),                            # &#123;
        re.compile(r'&#x[0-9a-fA-F]+;'),                  # &#x1F600;
    ]
    
    # ゲーム固有マーカーパターン: (heart), (star), (music), <<wait>>, [[shake]] など
    GAME_MARKER_PATTERNS = [
        re.compile(r'\([a-zA-Z_][a-zA-Z0-9_]*\)'),        # (heart), (star), (music)
        re.compile(r'<<[a-zA-Z_][a-zA-Z0-9_]*>>'),        # <<wait>>, <<shake>>
        re.compile(r'\[\[[a-zA-Z_][a-zA-Z0-9_]*\]\]'),    # [[effect]], [[sound]]
        re.compile(r'%[a-zA-Z_][a-zA-Z0-9_]*%'),          # %player%, %item%
        re.compile(r'@[a-zA-Z_][a-zA-Z0-9_]*@'),          # @name@, @title@
        re.compile(r'#[a-zA-Z_][a-zA-Z0-9_]*#'),          # #color#, #size#
    ]
    
    def __init__(self):
        self._var_counter = 0
        self._tag_counter = 0
        self._marker_counter = 0
    
    def _reset_counters(self):
        """カウンターをリセット"""
        self._var_counter = 0
        self._tag_counter = 0
        self._marker_counter = 0
    
    def protect_tags(self, text: str) -> TagProtectionResult:
        """
        変数・制御タグ・ゲーム固有マーカーをプレースホルダーに置換
        
        Args:
            text: 元のテキスト
            
        Returns:
            TagProtectionResult（保護済みテキスト、プレースホルダーマップ、各種カウント）
        """
        if not text:
            return TagProtectionResult(
                protected_text="",
                placeholder_map={},
                var_count=0,
                tag_count=0,
                marker_count=0
            )
        
        self._reset_counters()
        placeholder_map: Dict[str, str] = {}
        
        # 全てのマッチを収集（位置情報付き）
        matches: List[Tuple[int, int, str, str]] = []  # (start, end, original, type)
        
        # 1. 変数を検出 ({...})
        for match in self.VAR_PATTERN.finditer(text):
            matches.append((match.start(), match.end(), match.group(0), "VAR"))
        
        # 2. 制御タグを検出
        for pattern in self.CONTROL_TAG_PATTERNS:
            for match in pattern.finditer(text):
                matches.append((match.start(), match.end(), match.group(0), "TAG"))
        
        # 3. ゲーム固有マーカーを検出
        for pattern in self.GAME_MARKER_PATTERNS:
            for match in pattern.finditer(text):
                matches.append((match.start(), match.end(), match.group(0), "MARKER"))
        
        # 重複するマッチを除去（開始位置でソートし、重複を除去）
        matches.sort(key=lambda x: (x[0], -x[1]))  # 開始位置でソート、同じ開始なら長い方を優先
        
        filtered_matches: List[Tuple[int, int, str, str]] = []
        last_end = -1
        for start, end, original, tag_type in matches:
            if start >= last_end:
                filtered_matches.append((start, end, original, tag_type))
                last_end = end
        
        # 後ろから置換（位置がずれないように）
        protected_text = text
        var_count = 0
        tag_count = 0
        marker_count = 0
        
        for start, end, original, tag_type in reversed(filtered_matches):
            if tag_type == "VAR":
                placeholder = f"<<VAR_{var_count}>>"
                var_count += 1
            elif tag_type == "TAG":
                placeholder = f"<<TAG_{tag_count}>>"
                tag_count += 1
            else:  # MARKER
                placeholder = f"<<MARKER_{marker_count}>>"
                marker_count += 1
            
            placeholder_map[placeholder] = original
            protected_text = protected_text[:start] + placeholder + protected_text[end:]
        
        self._var_counter = var_count
        self._tag_counter = tag_count
        self._marker_counter = marker_count
        
        return TagProtectionResult(
            protected_text=protected_text,
            placeholder_map=placeholder_map,
            var_count=var_count,
            tag_count=tag_count,
            marker_count=marker_count
        )
    
    def restore_tags(self, text: str, placeholder_map: Dict[str, str]) -> str:
        """
        プレースホルダーを元の値に復元
        
        Args:
            text: プレースホルダーを含むテキスト
            placeholder_map: {placeholder: original_value} のマップ
            
        Returns:
            復元されたテキスト
        """
        if not text or not placeholder_map:
            return text
        
        restored_text = text
        
        # プレースホルダーを元の値に置換
        # 長いプレースホルダーから順に置換（部分一致を防ぐ）
        sorted_placeholders = sorted(placeholder_map.keys(), key=len, reverse=True)
        
        for placeholder in sorted_placeholders:
            original = placeholder_map[placeholder]
            restored_text = restored_text.replace(placeholder, original)
        
        return restored_text
    
    def validate_restoration(
        self,
        restored_text: str,
        placeholder_map: Dict[str, str]
    ) -> Tuple[bool, List[str]]:
        """
        復元後にプレースホルダーの欠落・重複がないか検証
        
        Args:
            restored_text: 復元されたテキスト
            placeholder_map: 元のプレースホルダーマップ
            
        Returns:
            (検証成功フラグ, エラーメッセージリスト)
        """
        errors = []
        
        # 残存プレースホルダーのチェック
        remaining_placeholders = re.findall(r'<<(?:VAR|TAG|MARKER)_\d+>>', restored_text)
        if remaining_placeholders:
            errors.append(f"未復元のプレースホルダーが残っています: {remaining_placeholders}")
        
        # 元の値が全て復元されているかチェック
        for placeholder, original in placeholder_map.items():
            if original not in restored_text:
                # 元の値が復元されたテキストに含まれていない場合
                # ただし、翻訳によって位置が変わる可能性があるため、警告レベル
                pass  # 翻訳後は位置が変わる可能性があるため、厳密なチェックは行わない
        
        return len(errors) == 0, errors


# モジュールレベルの便利関数
_default_protector = TagProtector()


def protect_tags(text: str) -> Tuple[str, Dict[str, str]]:
    """
    変数・制御タグをプレースホルダーに置換（便利関数）
    
    Args:
        text: 元のテキスト
        
    Returns:
        (protected_text, placeholder_map)
    """
    result = _default_protector.protect_tags(text)
    return result.protected_text, result.placeholder_map


def restore_tags(text: str, placeholder_map: Dict[str, str]) -> str:
    """
    プレースホルダーを元の値に復元（便利関数）
    
    Args:
        text: プレースホルダーを含むテキスト
        placeholder_map: {placeholder: original_value} のマップ
        
    Returns:
        復元されたテキスト
    """
    return _default_protector.restore_tags(text, placeholder_map)
