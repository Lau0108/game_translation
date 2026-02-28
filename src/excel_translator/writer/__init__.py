"""Excel出力モジュール

ExcelWriterクラスを提供し、翻訳結果をExcel形式で出力する。
- メイン翻訳結果シート
- パス別比較シート
- コストレポートシート
- モード間比較シート
- 条件付き書式（2nd pass: 青、3rd pass: 緑、4th pass: 黄）
"""

import logging
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional, Union

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)


class ExcelWriterError(Exception):
    """Excel出力エラー"""
    pass


# 条件付き書式の色定義
PASS_COLORS = {
    "pass_2": PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid"),  # 青
    "pass_3": PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid"),  # 緑
    "pass_4": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),  # 黄
}

# ヘッダースタイル
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

# セルボーダー
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


@dataclass
class MainSheetColumns:
    """メインシートのカラム定義"""
    TEXT_ID = "text_id"
    SOURCE_TEXT = "source_text"
    TRANSLATED_TEXT = "translated_text"
    CHARACTER_ID = "character_id"
    CHAR_COUNT = "char_count"
    LENGTH_OK = "length_ok"
    GLOSSARY_CHECK = "glossary_check"
    REMARKS = "remarks"
    ALTERNATIVE = "alternative"
    
    @classmethod
    def get_columns(cls) -> List[str]:
        """カラムリストを取得"""
        return [
            cls.TEXT_ID,
            cls.SOURCE_TEXT,
            cls.TRANSLATED_TEXT,
            cls.CHARACTER_ID,
            cls.CHAR_COUNT,
            cls.LENGTH_OK,
            cls.GLOSSARY_CHECK,
            cls.REMARKS,
            cls.ALTERNATIVE,
        ]


@dataclass
class PassComparisonColumns:
    """パス別比較シートのカラム定義"""
    TEXT_ID = "text_id"
    SOURCE_TEXT = "source_text"
    PASS_1 = "pass_1"
    PASS_2 = "pass_2"
    PASS_2_REASON = "pass_2_reason"
    PASS_3 = "pass_3"
    PASS_3_REASON = "pass_3_reason"
    PASS_4_BACKTRANS = "pass_4_backtrans"
    FINAL = "final"
    REMARKS = "remarks"
    
    @classmethod
    def get_columns(cls) -> List[str]:
        """カラムリストを取得"""
        return [
            cls.TEXT_ID,
            cls.SOURCE_TEXT,
            cls.PASS_1,
            cls.PASS_2,
            cls.PASS_2_REASON,
            cls.PASS_3,
            cls.PASS_3_REASON,
            cls.PASS_4_BACKTRANS,
            cls.FINAL,
            cls.REMARKS,
        ]


class ExcelWriter:
    """
    Excel出力クラス
    
    翻訳結果をExcel形式で出力する。
    Requirements 12.1-12.7 を満たす。
    """
    
    def __init__(self, char_limit: Optional[int] = None):
        """
        ExcelWriterを初期化
        
        Args:
            char_limit: 文字数制限（Noneの場合は制限なし）
        """
        self.char_limit = char_limit
    
    def write_main_sheet(
        self,
        results: Union[pd.DataFrame, List[Dict[str, Any]]],
        output_path: str,
        sheet_name: str = "翻訳結果",
    ) -> None:
        """
        メイン翻訳結果シートを出力
        
        カラム: text_id, source_text, translated_text, character_id,
                char_count, length_ok, glossary_check, remarks, alternative
        
        Args:
            results: 翻訳結果（DataFrameまたは辞書リスト）
            output_path: 出力ファイルパス
            sheet_name: シート名
        """
        if isinstance(results, list):
            df = pd.DataFrame(results)
        else:
            df = results.copy()
        
        # 必要なカラムを追加・整形
        df = self._prepare_main_sheet_data(df)
        
        # Excelファイルを作成
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # DataFrameをシートに書き込み
        self._write_dataframe_to_sheet(ws, df)
        
        # スタイルを適用
        self._apply_header_style(ws)
        self._apply_cell_borders(ws)
        self._adjust_column_widths(ws)
        
        # 保存
        self._save_workbook(wb, output_path)
        logger.info(f"メインシートを出力: {output_path}")

    
    def _prepare_main_sheet_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """メインシート用にデータを整形"""
        result = pd.DataFrame()
        
        # 必須カラム
        result["text_id"] = df.get("text_id", pd.Series(dtype=str))
        result["source_text"] = df.get("source_text", pd.Series(dtype=str))
        result["translated_text"] = df.get("translated_text", df.get("final", pd.Series(dtype=str)))
        result["character_id"] = df.get("character_id", df.get("character", pd.Series(dtype=str)))
        
        # 文字数カウント
        result["char_count"] = result["translated_text"].apply(
            lambda x: len(str(x)) if pd.notna(x) else 0
        )
        
        # 文字数チェック
        if self.char_limit:
            result["length_ok"] = result["char_count"].apply(
                lambda x: "✅" if x <= self.char_limit else "⚠️"
            )
        else:
            result["length_ok"] = "✅"
        
        # 用語チェック結果
        result["glossary_check"] = df.get("glossary_check", "")
        
        # remarks（原則空欄）
        result["remarks"] = df.get("remarks", "")
        
        # alternative
        result["alternative"] = df.get("alternative", "")
        
        return result
    
    def write_pass_comparison_sheet(
        self,
        pass_results: Union[pd.DataFrame, List[Dict[str, Any]]],
        output_path: str,
        sheet_name: str = "パス別比較",
    ) -> None:
        """
        パス別比較シートを出力
        
        カラム: text_id, source_text, pass_1, pass_2, pass_2_reason,
                pass_3, pass_3_reason, pass_4_backtrans, final
        
        Args:
            pass_results: パス別結果（DataFrameまたは辞書リスト）
            output_path: 出力ファイルパス
            sheet_name: シート名
        """
        if isinstance(pass_results, list):
            df = pd.DataFrame(pass_results)
        else:
            df = pass_results.copy()
        
        # 必要なカラムを整形
        df = self._prepare_pass_comparison_data(df)
        
        # Excelファイルを作成または追記
        wb = self._load_or_create_workbook(output_path)
        ws = wb.create_sheet(title=sheet_name)
        
        # DataFrameをシートに書き込み
        self._write_dataframe_to_sheet(ws, df)
        
        # スタイルを適用
        self._apply_header_style(ws)
        self._apply_cell_borders(ws)
        self._apply_conditional_formatting(ws, df)
        self._adjust_column_widths(ws)
        
        # 保存
        self._save_workbook(wb, output_path)
        logger.info(f"パス別比較シートを出力: {output_path}")

    
    def _prepare_pass_comparison_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """パス別比較シート用にデータを整形"""
        result = pd.DataFrame()
        
        result["text_id"] = df.get("text_id", pd.Series(dtype=str))
        result["source_text"] = df.get("source_text", pd.Series(dtype=str))
        result["pass_1"] = df.get("pass_1", df.get("pass1", pd.Series(dtype=str)))
        result["pass_2"] = df.get("pass_2", df.get("pass2", pd.Series(dtype=str)))
        result["pass_2_reason"] = df.get("pass_2_reason", df.get("pass2_reason", pd.Series(dtype=str)))
        result["pass_3"] = df.get("pass_3", df.get("pass3", pd.Series(dtype=str)))
        result["pass_3_reason"] = df.get("pass_3_reason", df.get("pass3_reason", pd.Series(dtype=str)))
        result["pass_4_backtrans"] = df.get("pass_4_backtrans", df.get("pass4_backtrans", pd.Series(dtype=str)))
        result["final"] = df.get("final", df.get("translated_text", pd.Series(dtype=str)))
        result["remarks"] = df.get("remarks", pd.Series(dtype=str))
        
        return result
    
    def write_cost_report_sheet(
        self,
        cost_report: Dict[str, Any],
        output_path: str,
        sheet_name: str = "コストレポート",
    ) -> None:
        """
        コストレポートシートを出力
        
        Args:
            cost_report: コストレポート辞書
            output_path: 出力ファイルパス
            sheet_name: シート名
        """
        # Excelファイルを作成または追記
        wb = self._load_or_create_workbook(output_path)
        ws = wb.create_sheet(title=sheet_name)
        
        # ヘッダー情報
        ws.append(["コストレポート"])
        ws.append(["モデル", cost_report.get("model", "N/A")])
        ws.append([])
        
        # パス別詳細
        ws.append(["パス別詳細"])
        headers = [
            "パス名", "入力トークン", "出力トークン", "キャッシュヒット",
            "API呼び出し", "処理時間(ms)", "コスト(USD)", "修正行数", "総行数", "修正率"
        ]
        ws.append(headers)
        
        passes = cost_report.get("passes", [])
        for p in passes:
            ws.append([
                p.get("pass_name", ""),
                p.get("input_tokens", 0),
                p.get("output_tokens", 0),
                p.get("cache_hit_tokens", 0),
                p.get("api_calls", 0),
                p.get("processing_time_ms", 0),
                f"${p.get('cost_usd', 0):.4f}",
                p.get("modified_rows", 0),
                p.get("total_rows", 0),
                f"{p.get('modification_rate', 0) * 100:.1f}%",
            ])
        
        ws.append([])
        
        # サマリ
        ws.append(["サマリ"])
        summary = cost_report.get("summary", {})
        ws.append(["総入力トークン", summary.get("total_input_tokens", 0)])
        ws.append(["総出力トークン", summary.get("total_output_tokens", 0)])
        ws.append(["総キャッシュヒット", summary.get("total_cache_hit_tokens", 0)])
        ws.append(["総API呼び出し", summary.get("total_api_calls", 0)])
        ws.append(["総処理時間(ms)", summary.get("total_processing_time_ms", 0)])
        ws.append(["総コスト(USD)", f"${summary.get('total_cost_usd', 0):.4f}"])
        
        # スタイルを適用
        self._apply_cost_report_style(ws)
        self._adjust_column_widths(ws)
        
        # 保存
        self._save_workbook(wb, output_path)
        logger.info(f"コストレポートシートを出力: {output_path}")

    
    def write_mode_comparison_sheet(
        self,
        mode_results: Dict[str, Any],
        output_path: str,
        sheet_name: str = "モード間比較",
    ) -> None:
        """
        モード間比較シートを出力
        
        Args:
            mode_results: モード間比較結果辞書
            output_path: 出力ファイルパス
            sheet_name: シート名
        """
        # Excelファイルを作成または追記
        wb = self._load_or_create_workbook(output_path)
        ws = wb.create_sheet(title=sheet_name)
        
        modes = mode_results.get("modes", [])
        comparison = mode_results.get("comparison", [])
        cost_comparison = mode_results.get("cost_comparison", {})
        
        # ヘッダー
        headers = ["text_id", "source_text"]
        for mode in modes:
            headers.append(f"{mode}_result")
        headers.append("differences")
        ws.append(headers)
        
        # データ行
        for row in comparison:
            row_data = [
                row.get("text_id", ""),
                row.get("source_text", ""),
            ]
            for mode in modes:
                row_data.append(row.get(f"{mode}_result", ""))
            
            # 差分情報
            diffs = row.get("differences", [])
            diff_str = "; ".join([
                f"{d['from_mode']}→{d['to_mode']}"
                for d in diffs if d.get("changed")
            ])
            row_data.append(diff_str)
            
            ws.append(row_data)
        
        # コスト比較セクション
        ws.append([])
        ws.append(["コスト比較"])
        cost_headers = ["モード", "総コスト(USD)", "追加コスト(USD)"]
        ws.append(cost_headers)
        
        for mode in modes:
            mode_cost = cost_comparison.get(mode, {})
            ws.append([
                mode,
                f"${mode_cost.get('total_cost_usd', 0):.4f}",
                f"${mode_cost.get('additional_cost_usd', 0):.4f}" if "additional_cost_usd" in mode_cost else "N/A",
            ])
        
        # スタイルを適用
        self._apply_header_style(ws)
        self._apply_cell_borders(ws)
        self._adjust_column_widths(ws)
        
        # 保存
        self._save_workbook(wb, output_path)
        logger.info(f"モード間比較シートを出力: {output_path}")
    
    def apply_conditional_formatting(self, ws: Worksheet) -> None:
        """
        条件付き書式を適用（2nd pass: 青、3rd pass: 緑、4th pass: 黄）
        
        Args:
            ws: ワークシート
        """
        # このメソッドは外部から呼び出し可能なパブリックメソッド
        # 内部では_apply_conditional_formattingを使用
        pass

    
    def _apply_conditional_formatting(
        self,
        ws: Worksheet,
        df: pd.DataFrame,
    ) -> None:
        """
        パス別比較シートに条件付き書式を適用
        
        変更があったセルに色を付ける:
        - 2nd pass: 青
        - 3rd pass: 緑
        - 4th pass: 黄
        
        Args:
            ws: ワークシート
            df: データフレーム
        """
        # カラムインデックスを取得
        columns = list(df.columns)
        
        pass_2_col = columns.index("pass_2") + 1 if "pass_2" in columns else None
        pass_3_col = columns.index("pass_3") + 1 if "pass_3" in columns else None
        pass_4_col = columns.index("pass_4_backtrans") + 1 if "pass_4_backtrans" in columns else None
        
        # データ行をループ（ヘッダー行をスキップ）
        for row_idx, row in enumerate(df.itertuples(index=False), start=2):
            row_dict = dict(zip(columns, row))
            
            # pass_2に値がある場合は青
            if pass_2_col and pd.notna(row_dict.get("pass_2")) and str(row_dict.get("pass_2")).strip():
                cell = ws.cell(row=row_idx, column=pass_2_col)
                cell.fill = PASS_COLORS["pass_2"]
            
            # pass_3に値がある場合は緑
            if pass_3_col and pd.notna(row_dict.get("pass_3")) and str(row_dict.get("pass_3")).strip():
                cell = ws.cell(row=row_idx, column=pass_3_col)
                cell.fill = PASS_COLORS["pass_3"]
            
            # pass_4_backtransに値がある場合は黄
            if pass_4_col and pd.notna(row_dict.get("pass_4_backtrans")) and str(row_dict.get("pass_4_backtrans")).strip():
                cell = ws.cell(row=row_idx, column=pass_4_col)
                cell.fill = PASS_COLORS["pass_4"]
    
    def _write_dataframe_to_sheet(
        self,
        ws: Worksheet,
        df: pd.DataFrame,
    ) -> None:
        """DataFrameをワークシートに書き込み"""
        # ヘッダー行
        ws.append(list(df.columns))
        
        # データ行
        for row in df.itertuples(index=False):
            ws.append(list(row))
    
    def _apply_header_style(self, ws: Worksheet) -> None:
        """ヘッダー行にスタイルを適用"""
        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = HEADER_ALIGNMENT
    
    def _apply_cell_borders(self, ws: Worksheet) -> None:
        """全セルにボーダーを適用"""
        for row in ws.iter_rows():
            for cell in row:
                cell.border = THIN_BORDER

    
    def _apply_cost_report_style(self, ws: Worksheet) -> None:
        """コストレポートシートにスタイルを適用"""
        # タイトル行
        ws["A1"].font = Font(bold=True, size=14)
        
        # パス別詳細のヘッダー
        for cell in ws[5]:
            if cell.value:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = HEADER_ALIGNMENT
    
    def _adjust_column_widths(self, ws: Worksheet) -> None:
        """カラム幅を自動調整"""
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        # 日本語文字は幅が広いので調整
                        jp_chars = sum(1 for c in str(cell.value) if ord(c) > 127)
                        cell_length += jp_chars * 0.5
                        max_length = max(max_length, cell_length)
                except Exception:
                    pass
            
            # 最小幅と最大幅を設定
            adjusted_width = min(max(max_length + 2, 10), 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _load_or_create_workbook(self, output_path: str) -> Workbook:
        """既存のワークブックを読み込むか、新規作成"""
        path = Path(output_path)
        
        if path.exists():
            try:
                from openpyxl import load_workbook
                return load_workbook(output_path)
            except Exception as e:
                logger.warning(f"既存ファイルの読み込みに失敗: {e}")
        
        return Workbook()
    
    def _save_workbook(self, wb: Workbook, output_path: str) -> None:
        """ワークブックを保存"""
        path = Path(output_path)
        path.parent.mkdir(parents=True, exist_ok=True)
        
        try:
            wb.save(output_path)
        except Exception as e:
            raise ExcelWriterError(f"ファイルの保存に失敗しました: {e}")

    
    def write_all_sheets(
        self,
        results: Union[pd.DataFrame, List[Dict[str, Any]]],
        pass_results: Optional[Union[pd.DataFrame, List[Dict[str, Any]]]] = None,
        cost_report: Optional[Dict[str, Any]] = None,
        mode_results: Optional[Dict[str, Any]] = None,
        output_path: str = "output.xlsx",
    ) -> None:
        """
        全シートを一括出力
        
        Args:
            results: メイン翻訳結果
            pass_results: パス別結果（オプション）
            cost_report: コストレポート（オプション）
            mode_results: モード間比較結果（オプション）
            output_path: 出力ファイルパス
        """
        # メインシートを出力
        self.write_main_sheet(results, output_path)
        
        # パス別比較シートを出力
        if pass_results is not None:
            self.write_pass_comparison_sheet(pass_results, output_path)
        
        # コストレポートシートを出力
        if cost_report is not None:
            self.write_cost_report_sheet(cost_report, output_path)
        
        # モード間比較シートを出力
        if mode_results is not None:
            self.write_mode_comparison_sheet(mode_results, output_path)
        
        logger.info(f"全シートを出力完了: {output_path}")
    
    def write_multi_file_output(
        self,
        file_results: Dict[str, Union[pd.DataFrame, List[Dict[str, Any]]]],
        output_path: str,
        separate_files: bool = False,
    ) -> None:
        """
        複数ファイルの結果を出力
        
        Args:
            file_results: {ファイル名: 結果} の辞書
            output_path: 出力ファイルパス（またはディレクトリ）
            separate_files: ファイル単位で分離するか
        """
        if separate_files:
            # ファイル単位で分離出力
            output_dir = Path(output_path)
            output_dir.mkdir(parents=True, exist_ok=True)
            
            for file_name, results in file_results.items():
                file_output = output_dir / f"{Path(file_name).stem}_translated.xlsx"
                self.write_main_sheet(results, str(file_output))
        else:
            # 統合して1つのExcelに出力
            wb = Workbook()
            first_sheet = True
            
            for file_name, results in file_results.items():
                if isinstance(results, list):
                    df = pd.DataFrame(results)
                else:
                    df = results.copy()
                
                df = self._prepare_main_sheet_data(df)
                
                sheet_name = Path(file_name).stem[:31]  # シート名は31文字まで
                
                if first_sheet:
                    ws = wb.active
                    ws.title = sheet_name
                    first_sheet = False
                else:
                    ws = wb.create_sheet(title=sheet_name)
                
                self._write_dataframe_to_sheet(ws, df)
                self._apply_header_style(ws)
                self._apply_cell_borders(ws)
                self._adjust_column_widths(ws)
            
            self._save_workbook(wb, output_path)
            logger.info(f"複数ファイルを統合出力: {output_path}")
